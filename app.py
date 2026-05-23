"""
Quản Lý Thu Chi Căn Hộ — v4.1 (UI Refresh)
Database: Google Sheets (dữ liệu vĩnh viễn, không mất khi deploy lại)
TT152/2025/TT-BTC

Nâng cấp v4.1:
- UI/UX nâng cấp toàn diện: dashboard cards màu sắc, form compact, toast notifications
- Giữ nguyên 100% logic: @st.fragment, Google Sheets sync, tính toán dòng tiền
"""

import streamlit as st
import pandas as pd
import gspread
from google.oauth2.service_account import Credentials
import json
import io
import calendar
from datetime import date, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ══════════════════════════════════════════════════════════════════════════════
# CONSTANTS
# ══════════════════════════════════════════════════════════════════════════════
ROOMS = ['P101', 'P201', 'P202', 'P301', 'P302', 'P401', 'P402', 'P501', 'P502']

DANH_MUC_THU = [
    "Doanh thu tiền phòng",
    "Thu hộ tiền điện",
    "Thu hộ tiền nước",
]

DANH_MUC_CHI = [
    "Chi phí vận hành chung",
    "Chi lương / thưởng",
    "Chi quản lý chung",
    "Chi hộ tiền điện",
    "Chi hộ tiền nước",
    "Chi khác",
]

KY_OPTIONS = [
    "Tháng này", "Quý 1", "Quý 2", "Quý 3", "Quý 4",
    "6 Tháng đầu năm", "6 Tháng cuối năm", "Cả năm", "Tùy chỉnh khoảng ngày",
]

GS_SHEET_NAME = "giao_dich"
GS_COLUMNS = ["id", "ngay", "loai", "ma_phong", "danh_muc", "so_tien", "ghi_chu", "thang_nhap", "trang_thai"]
TRANG_THAI_DA_THU = {"Đã thu", "Đã thanh toán"}

# ── Tiền cọc tab ─────────────────────────────────────────────────────────────
GS_COC_SHEET   = "tiem_coc"
GS_COC_COLUMNS = ["ma_phong", "ten_kh", "ngay_coc", "so_tien_coc", "ghi_chu"]

# ══════════════════════════════════════════════════════════════════════════════
# GOOGLE SHEETS CONNECTION
# ══════════════════════════════════════════════════════════════════════════════
SCOPES = [
    "https://www.googleapis.com/auth/spreadsheets",
    "https://www.googleapis.com/auth/drive",
]

@st.cache_resource(ttl=300)
def get_gsheet_client():
    if "gcp_service_account" not in st.secrets:
        st.error(
            "❌ Thiếu `[gcp_service_account]` trong secrets.toml\n\n"
            "**Cách sửa:** Mở file `.streamlit/secrets.toml` và đảm bảo có đúng format bên dưới."
        )
        return None
    if "spreadsheet_url" not in st.secrets:
        st.error("❌ Thiếu `spreadsheet_url` trong secrets.toml")
        return None
    try:
        raw = st.secrets["gcp_service_account"]
        creds_dict = {
            "type":                        raw.get("type", "service_account"),
            "project_id":                  raw["project_id"],
            "private_key_id":              raw["private_key_id"],
            "private_key":                 raw["private_key"].replace("\\n", "\n"),
            "client_email":                raw["client_email"],
            "client_id":                   raw["client_id"],
            "auth_uri":                    raw.get("auth_uri", "https://accounts.google.com/o/oauth2/auth"),
            "token_uri":                   raw.get("token_uri", "https://oauth2.googleapis.com/token"),
            "auth_provider_x509_cert_url": raw.get("auth_provider_x509_cert_url", "https://www.googleapis.com/oauth2/v1/certs"),
            "client_x509_cert_url":        raw.get("client_x509_cert_url", ""),
        }
        creds  = Credentials.from_service_account_info(creds_dict, scopes=SCOPES)
        client = gspread.authorize(creds)
        return client
    except KeyError as e:
        st.error(f"❌ Thiếu trường `{e}` trong `[gcp_service_account]` của secrets.toml")
        return None
    except Exception as e:
        st.error(f"❌ Lỗi kết nối Google Sheets: {e}")
        return None

@st.cache_resource(ttl=300)
def get_worksheet():
    client = get_gsheet_client()
    if client is None:
        return None
    try:
        spreadsheet_url = st.secrets["spreadsheet_url"]
        sh = client.open_by_url(spreadsheet_url)
        try:
            ws = sh.worksheet(GS_SHEET_NAME)
        except gspread.WorksheetNotFound:
            ws = sh.add_worksheet(title=GS_SHEET_NAME, rows=5000, cols=len(GS_COLUMNS))
            ws.append_row(GS_COLUMNS)
        return ws
    except gspread.exceptions.APIError as e:
        st.error(f"❌ Google Sheets API lỗi: {e}.")
        return None
    except Exception as e:
        st.error(f"❌ Không mở được Google Sheets: {e}")
        return None

@st.cache_resource(ttl=300)
def get_worksheet_coc():
    """Worksheet riêng cho tab tiem_coc — hoàn toàn tách biệt khỏi giao_dich."""
    client = get_gsheet_client()
    if client is None:
        return None
    try:
        spreadsheet_url = st.secrets["spreadsheet_url"]
        sh = client.open_by_url(spreadsheet_url)
        try:
            ws = sh.worksheet(GS_COC_SHEET)
        except gspread.WorksheetNotFound:
            ws = sh.add_worksheet(title=GS_COC_SHEET, rows=20, cols=len(GS_COC_COLUMNS))
            ws.append_row(GS_COC_COLUMNS)
            # Khởi tạo 9 dòng trống cho 9 phòng
            for room in ROOMS:
                ws.append_row([room, "", "", 0, ""])
        return ws
    except gspread.exceptions.APIError as e:
        st.error(f"❌ Google Sheets API lỗi (cọc): {e}.")
        return None
    except Exception as e:
        st.error(f"❌ Không mở được sheet tiền cọc: {e}")
        return None

# ══════════════════════════════════════════════════════════════════════════════
# DANH_SACH_PHONG — SHEET MỚI (Bước 1: Kết nối & Đọc dữ liệu phòng)
# ══════════════════════════════════════════════════════════════════════════════
GS_DSP_SHEET   = "DANH_SACH_PHONG"
GS_DSP_COLUMNS = [
    "Mã Phòng", "Tên Khách Thuê", "Giá Thuê Gốc",
    "Đơn Giá Điện", "Đơn Giá Nước", "Ngôn Ngữ Bill",
]

# Dữ liệu mặc định khởi tạo cho 9 phòng khi tạo sheet lần đầu
_DSP_DEFAULT_LANG = "Tiếng Việt"
_DSP_SEED_ROWS = [
    [room, "", 0, 0, 0, _DSP_DEFAULT_LANG] for room in ROOMS
]

@st.cache_resource(ttl=300)
def get_worksheet_dsp():
    """
    Trả về worksheet 'DANH_SACH_PHONG'.
    Nếu sheet chưa tồn tại → tự tạo mới và seed 9 dòng cho 9 phòng.
    Hoàn toàn độc lập, KHÔNG ảnh hưởng đến giao_dich / tiem_coc.
    """
    client = get_gsheet_client()
    if client is None:
        return None
    try:
        spreadsheet_url = st.secrets["spreadsheet_url"]
        sh = client.open_by_url(spreadsheet_url)
        try:
            ws = sh.worksheet(GS_DSP_SHEET)
        except gspread.WorksheetNotFound:
            ws = sh.add_worksheet(
                title=GS_DSP_SHEET,
                rows=len(ROOMS) + 10,
                cols=len(GS_DSP_COLUMNS),
            )
            # Header row
            ws.append_row(GS_DSP_COLUMNS)
            # Seed 9 dòng — 1 dòng mỗi phòng, giá trị mặc định = 0
            ws.append_rows(_DSP_SEED_ROWS, value_input_option="USER_ENTERED")
        return ws
    except gspread.exceptions.APIError as e:
        st.error(f"❌ Google Sheets API lỗi (danh sách phòng): {e}.")
        return None
    except Exception as e:
        st.error(f"❌ Không mở được sheet danh sách phòng: {e}")
        return None


@st.cache_data(ttl=60, show_spinner=False)
def load_danh_sach_phong() -> pd.DataFrame:
    """
    Đọc toàn bộ sheet DANH_SACH_PHONG thành DataFrame.
    Cache 60 giây — gọi load_danh_sach_phong.clear() sau mỗi lần ghi để làm mới.
    """
    ws = get_worksheet_dsp()
    if ws is None:
        return pd.DataFrame(columns=GS_DSP_COLUMNS)
    try:
        data = ws.get_all_records(expected_headers=GS_DSP_COLUMNS)
        if not data:
            return pd.DataFrame(columns=GS_DSP_COLUMNS)
        df = pd.DataFrame(data)
        df["Giá Thuê Gốc"]  = pd.to_numeric(df["Giá Thuê Gốc"],  errors="coerce").fillna(0).astype(int)
        df["Đơn Giá Điện"]  = pd.to_numeric(df["Đơn Giá Điện"],  errors="coerce").fillna(0).astype(int)
        df["Đơn Giá Nước"]  = pd.to_numeric(df["Đơn Giá Nước"],  errors="coerce").fillna(0).astype(int)
        df["Ngôn Ngữ Bill"] = df["Ngôn Ngữ Bill"].fillna(_DSP_DEFAULT_LANG).astype(str)
        # Đảm bảo đủ 9 phòng (phòng chưa có dòng → thêm dòng mặc định)
        existing = set(df["Mã Phòng"].tolist())
        missing  = [r for r in ROOMS if r not in existing]
        if missing:
            empty = pd.DataFrame([{
                "Mã Phòng": r, "Tên Khách Thuê": "", "Giá Thuê Gốc": 0,
                "Đơn Giá Điện": 0, "Đơn Giá Nước": 0,
                "Ngôn Ngữ Bill": _DSP_DEFAULT_LANG,
            } for r in missing])
            df = pd.concat([df, empty], ignore_index=True)
        # Sort theo thứ tự ROOMS
        room_order = {r: i for i, r in enumerate(ROOMS)}
        df["_order"] = df["Mã Phòng"].map(room_order)
        df = df.sort_values("_order").drop(columns=["_order"]).reset_index(drop=True)
        return df
    except Exception as e:
        st.error(f"Lỗi đọc danh sách phòng: {e}")
        return pd.DataFrame(columns=GS_DSP_COLUMNS)


def get_thong_tin_phong(ma_phong: str) -> dict:
    """
    Trả về dict chứa toàn bộ thông tin gốc của phòng từ sheet DANH_SACH_PHONG.

    Ví dụ kết quả:
    {
        "Mã Phòng":       "P101",
        "Tên Khách Thuê": "Nguyễn Văn A",
        "Giá Thuê Gốc":   5000000,
        "Đơn Giá Điện":   3500,
        "Đơn Giá Nước":   15000,
        "Ngôn Ngữ Bill":  "Tiếng Việt",
    }

    Nếu phòng không tìm thấy → trả về dict với giá trị mặc định (không raise lỗi).
    """
    df = load_danh_sach_phong()
    row = df[df["Mã Phòng"] == ma_phong]
    if row.empty:
        return {
            "Mã Phòng":       ma_phong,
            "Tên Khách Thuê": "",
            "Giá Thuê Gốc":   0,
            "Đơn Giá Điện":   0,
            "Đơn Giá Nước":   0,
            "Ngôn Ngữ Bill":  _DSP_DEFAULT_LANG,
        }
    return row.iloc[0].to_dict()


def upsert_thong_tin_phong(
    ma_phong: str,
    ten_kh: str,
    gia_thue_goc: int,
    don_gia_dien: int,
    don_gia_nuoc: int,
    ngon_ngu_bill: str = "Tiếng Việt",
) -> bool:
    """
    Ghi đè (hoặc thêm mới) 1 dòng trong DANH_SACH_PHONG cho phòng ma_phong.
    Trả về True nếu thành công, False nếu lỗi.
    """
    ws = get_worksheet_dsp()
    if ws is None:
        return False
    try:
        all_rooms = ws.col_values(1)  # ['Mã Phòng', 'P101', 'P201', ...]
        target_row = None
        for idx, val in enumerate(all_rooms):
            if val == ma_phong:
                target_row = idx + 1  # 1-based
                break

        row_data = [ma_phong, ten_kh, gia_thue_goc, don_gia_dien, don_gia_nuoc, ngon_ngu_bill]

        if target_row:
            col_end   = gspread.utils.rowcol_to_a1(target_row, len(GS_DSP_COLUMNS))
            col_start = gspread.utils.rowcol_to_a1(target_row, 1)
            ws.update(f"{col_start}:{col_end}", [row_data], value_input_option="USER_ENTERED")
        else:
            ws.append_row(row_data, value_input_option="USER_ENTERED")

        load_danh_sach_phong.clear()
        return True
    except Exception as e:
        st.error(f"Lỗi lưu thông tin phòng: {e}")
        return False


# ══════════════════════════════════════════════════════════════════════════════
# DATABASE OPERATIONS
# ══════════════════════════════════════════════════════════════════════════════
@st.cache_data(ttl=60, show_spinner=False)
def load_all() -> pd.DataFrame:
    ws = get_worksheet()
    if ws is None:
        return pd.DataFrame(columns=GS_COLUMNS)
    try:
        data = ws.get_all_records(expected_headers=GS_COLUMNS)
        if not data:
            return pd.DataFrame(columns=GS_COLUMNS)
        df = pd.DataFrame(data)
        df["ngay"]    = pd.to_datetime(df["ngay"], errors="coerce")
        df["so_tien"] = pd.to_numeric(df["so_tien"], errors="coerce").fillna(0)
        df = df.dropna(subset=["ngay"])
        return df.sort_values("ngay", ascending=False).reset_index(drop=True)
    except Exception as e:
        st.error(f"Lỗi đọc dữ liệu: {e}")
        return pd.DataFrame(columns=GS_COLUMNS)

def parse_tien(s: str) -> int:
    s = str(s).strip()
    if not s:
        return 0
    s = s.replace(" ", "").replace("₫", "").replace("đ", "").replace("VND", "")
    if "." in s and "," in s:
        last_dot   = s.rfind(".")
        last_comma = s.rfind(",")
        if last_dot > last_comma:
            s = s.replace(",", "")
        else:
            s = s.replace(".", "").replace(",", ".")
    elif "." in s:
        parts = s.split(".")
        if all(len(p) == 3 for p in parts[1:]):
            s = s.replace(".", "")
    elif "," in s:
        parts = s.split(",")
        if all(len(p) == 3 for p in parts[1:]):
            s = s.replace(",", "")
        else:
            s = s.replace(",", ".")
    try:
        return max(0, int(float(s)))
    except (ValueError, TypeError):
        return 0

def _next_id(ws) -> int:
    try:
        data = ws.col_values(1)
        ids = [int(x) for x in data[1:] if str(x).isdigit()]
        return max(ids) + 1 if ids else 1
    except Exception:
        return 1

def insert_gd(ngay, loai, ma_phong, danh_muc, so_tien, ghi_chu, trang_thai="Đã thu"):
    ws = get_worksheet()
    if ws is None:
        return False
    try:
        new_id = _next_id(ws)
        thang_nhap = f"{ngay.month:02d}/{ngay.year}"
        row = [
            new_id, str(ngay), loai, ma_phong, danh_muc,
            float(so_tien), ghi_chu or "", thang_nhap, trang_thai,
        ]
        ws.append_row(row, value_input_option="USER_ENTERED")
        load_all.clear()
        return True
    except Exception as e:
        st.error(f"Lỗi lưu dữ liệu: {e}")
        return False

def append_batch(rows: list[list]) -> bool:
    ws = get_worksheet()
    if ws is None:
        return False
    try:
        next_id = _next_id(ws)
        full_rows = [[next_id + i] + r for i, r in enumerate(rows)]
        ws.append_rows(full_rows, value_input_option="USER_ENTERED")
        load_all.clear()
        return True
    except Exception as e:
        st.error(f"Lỗi ghi batch: {e}")
        return False

def delete_gd(gd_id: int):
    ws = get_worksheet()
    if ws is None:
        return
    try:
        cell = ws.find(str(gd_id), in_column=1)
        if cell:
            ws.delete_rows(cell.row)
            load_all.clear()
    except Exception as e:
        st.error(f"Lỗi xóa: {e}")

def update_gd(gd_id: int, ngay_moi, loai_moi: str, ma_phong_moi: str,
              danh_muc_moi: str, so_tien_moi: float, ghi_chu_moi: str) -> bool:
    ws = get_worksheet()
    if ws is None:
        return False
    try:
        cell = ws.find(str(gd_id), in_column=1)
        if not cell:
            st.error(f"Không tìm thấy giao dịch id={gd_id}")
            return False
        row_idx = cell.row
        thang_nhap = f"{ngay_moi.month:02d}/{ngay_moi.year}"
        updates = [
            {"range": gspread.utils.rowcol_to_a1(row_idx, 2), "values": [[str(ngay_moi)]]},
            {"range": gspread.utils.rowcol_to_a1(row_idx, 3), "values": [[loai_moi]]},
            {"range": gspread.utils.rowcol_to_a1(row_idx, 4), "values": [[ma_phong_moi]]},
            {"range": gspread.utils.rowcol_to_a1(row_idx, 5), "values": [[danh_muc_moi]]},
            {"range": gspread.utils.rowcol_to_a1(row_idx, 6), "values": [[float(so_tien_moi)]]},
            {"range": gspread.utils.rowcol_to_a1(row_idx, 7), "values": [[ghi_chu_moi or ""]]},
            {"range": gspread.utils.rowcol_to_a1(row_idx, 8), "values": [[thang_nhap]]},
        ]
        ws.batch_update(updates, value_input_option="USER_ENTERED")
        load_all.clear()
        return True
    except Exception as e:
        st.error(f"Lỗi sửa giao dịch: {e}")
        return False

def delete_batch(gd_ids: list[int]):
    ws = get_worksheet()
    if ws is None:
        return False
    try:
        all_ids = ws.col_values(1)
        rows_to_delete = []
        for row_idx, cell_val in enumerate(all_ids):
            try:
                if int(cell_val) in gd_ids:
                    rows_to_delete.append(row_idx + 1)
            except (ValueError, TypeError):
                pass
        for row_idx in sorted(rows_to_delete, reverse=True):
            ws.delete_rows(row_idx)
        load_all.clear()
        return True
    except Exception as e:
        st.error(f"Lỗi xóa batch: {e}")
        return False

def update_trang_thai(gd_ids: list[int], trang_thai_moi: str) -> bool:
    ws = get_worksheet()
    if ws is None:
        return False
    try:
        all_ids = ws.col_values(1)
        col_tt  = GS_COLUMNS.index("trang_thai") + 1

        updates = []
        for row_idx, cell_val in enumerate(all_ids):
            try:
                if int(cell_val) in gd_ids:
                    updates.append({
                        "range": gspread.utils.rowcol_to_a1(row_idx + 1, col_tt),
                        "values": [[trang_thai_moi]],
                    })
            except (ValueError, TypeError):
                pass

        if updates:
            ws.batch_update(updates, value_input_option="USER_ENTERED")
            load_all.clear()
        return True
    except Exception as e:
        st.error(f"Lỗi cập nhật trạng thái: {e}")
        return False

# ══════════════════════════════════════════════════════════════════════════════
# TIỀN CỌC — DATABASE OPERATIONS (hoàn toàn tách biệt khỏi thu chi)
# ══════════════════════════════════════════════════════════════════════════════
@st.cache_data(ttl=60, show_spinner=False)
def load_coc() -> pd.DataFrame:
    """Đọc toàn bộ dữ liệu tiền cọc từ sheet 'tiem_coc'."""
    ws = get_worksheet_coc()
    if ws is None:
        return pd.DataFrame(columns=GS_COC_COLUMNS)
    try:
        data = ws.get_all_records(expected_headers=GS_COC_COLUMNS)
        if not data:
            return pd.DataFrame(columns=GS_COC_COLUMNS)
        df = pd.DataFrame(data)
        df["so_tien_coc"] = pd.to_numeric(df["so_tien_coc"], errors="coerce").fillna(0)
        df["ngay_coc"]    = pd.to_datetime(df["ngay_coc"], errors="coerce")
        # Đảm bảo đủ 9 phòng
        existing_rooms = set(df["ma_phong"].tolist())
        missing = [r for r in ROOMS if r not in existing_rooms]
        if missing:
            empty_rows = pd.DataFrame([{
                "ma_phong": r, "ten_kh": "", "ngay_coc": pd.NaT,
                "so_tien_coc": 0, "ghi_chu": ""
            } for r in missing])
            df = pd.concat([df, empty_rows], ignore_index=True)
        # Sort theo thứ tự ROOMS
        room_order = {r: i for i, r in enumerate(ROOMS)}
        df["_order"] = df["ma_phong"].map(room_order)
        df = df.sort_values("_order").drop(columns=["_order"]).reset_index(drop=True)
        return df
    except Exception as e:
        st.error(f"Lỗi đọc dữ liệu tiền cọc: {e}")
        return pd.DataFrame(columns=GS_COC_COLUMNS)

def upsert_coc(ma_phong: str, ten_kh: str, ngay_coc, so_tien_coc: int, ghi_chu: str) -> bool:
    """Cập nhật (ghi đè) dòng cọc của phòng — dùng find() để locate row."""
    ws = get_worksheet_coc()
    if ws is None:
        return False
    try:
        # Tìm dòng theo ma_phong ở cột 1
        all_rooms = ws.col_values(1)  # ['ma_phong', 'P101', 'P201', ...]
        target_row = None
        for idx, val in enumerate(all_rooms):
            if val == ma_phong:
                target_row = idx + 1  # 1-based
                break

        ngay_str = str(ngay_coc) if ngay_coc else ""
        row_data  = [ma_phong, ten_kh, ngay_str, float(so_tien_coc), ghi_chu or ""]

        if target_row:
            # Ghi đè dòng hiện có (cột A→E)
            col_end = gspread.utils.rowcol_to_a1(target_row, len(GS_COC_COLUMNS))
            col_start = gspread.utils.rowcol_to_a1(target_row, 1)
            ws.update(f"{col_start}:{col_end}", [row_data], value_input_option="USER_ENTERED")
        else:
            # Phòng chưa có dòng → append mới
            ws.append_row(row_data, value_input_option="USER_ENTERED")

        load_coc.clear()
        return True
    except Exception as e:
        st.error(f"Lỗi lưu tiền cọc: {e}")
        return False

# ══════════════════════════════════════════════════════════════════════════════
# EXCEL IMPORT
# ══════════════════════════════════════════════════════════════════════════════
def parse_excel_invoice(uploaded_file, thang: int, nam: int) -> pd.DataFrame | None:
    try:
        raw = pd.read_excel(
            uploaded_file,
            sheet_name="3. QL hóa đơn",
            header=[1, 2],
            engine="openpyxl",
        )
    except Exception as e:
        st.error(f"❌ Không đọc được sheet '3. QL hóa đơn': {e}")
        return None

    def flatten_col(col_tuple):
        parts = [str(c).strip() for c in col_tuple if "Unnamed" not in str(c)]
        return "_".join(parts) if parts else ""

    flat_cols = [flatten_col(c) for c in raw.columns]
    raw.columns = flat_cols

    def find_col(keywords):
        for col in raw.columns:
            col_lower = col.lower()
            if all(k.lower() in col_lower for k in keywords):
                return col
        return None

    col_phong     = find_col(["phòng"])   or find_col(["phong"])
    col_dai_dien  = find_col(["đại diện"]) or find_col(["dai dien"])
    col_tong      = find_col(["tổng phải thu"]) or find_col(["tong phai thu"]) or find_col(["tổng"]) or find_col(["tong"])
    col_tien_nha  = (find_col(["tiền thuê phòng", "thành tiền"])
                     or find_col(["thuê phòng", "thành tiền"])
                     or find_col(["tiền thuê"])
                     or find_col(["tien thue"]))
    col_da_dung   = (find_col(["tiền điện", "đã dùng"])
                     or find_col(["tien dien", "da dung"])
                     or find_col(["đã dùng"])
                     or find_col(["da dung"]))
    col_tien_dien = (find_col(["tiền điện", "thành tiền"])
                     or find_col(["tien dien", "thanh tien"]))
    if col_tien_dien and col_da_dung and col_tien_dien == col_da_dung:
        col_tien_dien = None
    col_tien_nuoc = (find_col(["tiền nước"])
                     or find_col(["tien nuoc"]))

    found = {
        "Phòng":         col_phong,
        "Đại diện":      col_dai_dien,
        "Tiền nhà":      col_tien_nha,
        "Đã dùng (kWh)": col_da_dung,
        "Tiền điện":     col_tien_dien,
        "Tiền nước":     col_tien_nuoc,
        "Tổng phải thu": col_tong,
    }
    missing = [k for k, v in found.items() if v is None]
    if missing:
        st.warning(
            f"⚠️ Không tìm thấy cột: **{', '.join(missing)}**.\n\n"
            f"**Cột đọc được:** {list(raw.columns)}"
        )

    result_cols = {}
    if col_phong:     result_cols["ma_phong"]     = raw[col_phong].astype(str).str.strip()
    if col_dai_dien:  result_cols["ten_kh"]        = raw[col_dai_dien].astype(str).str.strip()
    if col_tien_nha:  result_cols["tien_phong"]    = pd.to_numeric(raw[col_tien_nha],  errors="coerce").fillna(0)
    if col_da_dung:   result_cols["so_dien_kwh"]   = pd.to_numeric(raw[col_da_dung],   errors="coerce").fillna(0)
    if col_tien_dien: result_cols["tien_dien"]     = pd.to_numeric(raw[col_tien_dien], errors="coerce").fillna(0)
    if col_tien_nuoc: result_cols["tien_nuoc"]     = pd.to_numeric(raw[col_tien_nuoc], errors="coerce").fillna(0)
    if col_tong:      result_cols["tong_phai_thu"] = pd.to_numeric(raw[col_tong],      errors="coerce").fillna(0)

    if not result_cols:
        st.error("❌ Không bóc tách được cột nào từ file Excel.")
        return None

    df_out = pd.DataFrame(result_cols)

    if "ma_phong" in df_out.columns:
        def _normalize_phong(x):
            try:
                s = str(x).strip()
                if s in ("nan", "None", ""):
                    return s
                if s.upper().startswith("P") and s[1:].isdigit():
                    return s.upper()
                num = str(int(float(s)))
                if num.isdigit():
                    return f"P{num}"
            except Exception:
                pass
            return s
        df_out["ma_phong"] = df_out["ma_phong"].apply(_normalize_phong)
        df_out = df_out[df_out["ma_phong"].isin(ROOMS)].copy()

    if df_out.empty:
        st.warning("⚠️ Không tìm thấy dữ liệu cho 9 phòng. Kiểm tra lại cột Phòng.")
        return None

    df_out["thang_nhap"] = f"{thang:02d}/{nam}"
    df_out["trang_thai"] = "Chờ thanh toán"
    return df_out.reset_index(drop=True)

# ══════════════════════════════════════════════════════════════════════════════
# HELPERS
# ══════════════════════════════════════════════════════════════════════════════
def fmt_vnd(val):
    try:
        return f"{int(val):,}".replace(",", ".") + " ₫"
    except Exception:
        return "0 ₫"

def get_date_range(ky: str, year: int, cs=None, ce=None):
    today = date.today()
    y = year
    if ky == "Tháng này":
        m = today.month
        s = date(y, m, 1)
        e = date(y, m, calendar.monthrange(y, m)[1])
    elif ky == "Quý 1":             s, e = date(y,1,1),  date(y,3,31)
    elif ky == "Quý 2":             s, e = date(y,4,1),  date(y,6,30)
    elif ky == "Quý 3":             s, e = date(y,7,1),  date(y,9,30)
    elif ky == "Quý 4":             s, e = date(y,10,1), date(y,12,31)
    elif ky == "6 Tháng đầu năm":  s, e = date(y,1,1),  date(y,6,30)
    elif ky == "6 Tháng cuối năm": s, e = date(y,7,1),  date(y,12,31)
    elif ky == "Cả năm":            s, e = date(y,1,1),  date(y,12,31)
    else:
        s = cs or date(y,1,1)
        e = ce or today
    return s, e

def compute_balances(df_all, s, e):
    if df_all.empty:
        return 0, 0, 0, 0
    df_real = df_all[df_all["trang_thai"].isin(TRANG_THAI_DA_THU)]
    if df_real.empty:
        return 0, 0, 0, 0
    prev = df_real[df_real["ngay"].dt.date < s]
    dau  = (prev[prev["loai"]=="Thu"]["so_tien"].sum()
          - prev[prev["loai"]=="Chi"]["so_tien"].sum())
    ky   = df_real[(df_real["ngay"].dt.date >= s) & (df_real["ngay"].dt.date <= e)]
    t    = ky[ky["loai"]=="Thu"]["so_tien"].sum()
    c    = ky[ky["loai"]=="Chi"]["so_tien"].sum()
    return dau, t, c, dau + t - c

# ══════════════════════════════════════════════════════════════════════════════
# EXCEL EXPORT
# ══════════════════════════════════════════════════════════════════════════════
def _bd():
    _s = Side(style="thin")
    return Border(left=_s, right=_s, top=_s, bottom=_s)

def _c(ws, row_n, c, v, bold=False, sz=10, col="000000", ha="center",
       fill=None, nf=None, wrap=False, italic=False):
    cell = ws.cell(row=row_n, column=c, value=v)
    cell.font      = Font(name="Arial", bold=bold, size=sz, color=col, italic=italic)
    cell.alignment = Alignment(horizontal=ha, vertical="center", wrap_text=wrap)
    cell.border    = _bd()
    if fill: cell.fill = PatternFill("solid", start_color=fill)
    if nf:   cell.number_format = nf
    return cell

def _mrow(ws, row_n, c1, c2, v, bold=False, sz=10, col="000000", ha="center", fill=None, italic=False):
    ws.merge_cells(start_row=row_n, start_column=c1, end_row=row_n, end_column=c2)
    cell = ws.cell(row=row_n, column=c1, value=v)
    cell.font      = Font(name="Arial", bold=bold, size=sz, color=col, italic=italic)
    cell.alignment = Alignment(horizontal=ha, vertical="center", wrap_text=True)
    if fill: cell.fill = PatternFill("solid", start_color=fill)
    return cell

def xuat_dong_tien(df_all, s, e) -> bytes:
    rows = []
    for _, row in df_all.iterrows():
        if (row["ngay"].date() >= s and row["ngay"].date() <= e):
            rows.append({
                "ngay": row["ngay"],
                "loai": str(row.get("loai", "")),
                "danh_muc": str(row.get("danh_muc", "")),
                "ma_phong": str(row.get("ma_phong", "")),
                "ghi_chu": str(row.get("ghi_chu", "") or ""),
                "so_tien": int(row["so_tien"]) if isinstance(row["so_tien"], (int, float)) else 0,
            })
    rows.sort(key=lambda x: x["ngay"])

    tong_thu = int(sum(r["so_tien"] for r in rows if r["loai"] == "Thu"))
    tong_chi = int(sum(r["so_tien"] for r in rows if r["loai"] == "Chi"))
    so_du = tong_thu - tong_chi

    wb = Workbook(); ws = wb.active; ws.title = "Dòng Tiền"
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 8
    ws.column_dimensions["C"].width = 28
    ws.column_dimensions["D"].width = 36
    ws.column_dimensions["E"].width = 22

    HDR_GREEN = "1a7f5a"; HDR_WHITE = "FFFFFF"
    ws.row_dimensions[1].height = 28
    _mrow(ws,1,1,5,"BÁO CÁO DÒNG TIỀN THỰC TẾ", bold=True, sz=14, ha="center", col=HDR_GREEN)
    _mrow(ws,2,1,5,f"Từ ngày {s.strftime('%d/%m/%Y')} đến ngày {e.strftime('%d/%m/%Y')}", sz=10, ha="center", italic=True)

    ws.row_dimensions[3].height = 18
    _c(ws,3,1,"Tổng thu", bold=True, sz=10, ha="left", col="1a5c3a", fill="d6f0e5")
    _c(ws,3,2,tong_thu, bold=True, sz=10, ha="right", nf="#,##0", fill="d6f0e5")
    _c(ws,3,3,"Tổng chi", bold=True, sz=10, ha="left", col="b52b2b", fill="fde8e8")
    _c(ws,3,4,tong_chi, bold=True, sz=10, ha="right", nf="#,##0", fill="fde8e8")
    _c(ws,3,5,"", fill="fff3cd")
    ws.row_dimensions[4].height = 22
    _mrow(ws,4,1,4, f"Số dư cuối kỳ: {'+' if so_du >= 0 else ''}{so_du:,} ₫",
          bold=True, sz=12, ha="right",
          col="1a5c3a" if so_du >= 0 else "b52b2b",
          fill="eaf7f1" if so_du >= 0 else "fdecea")

    ws.row_dimensions[5].height = 28
    for col, title in enumerate(["Ngày", "Loại", "Danh mục", "Diễn giải", "Số tiền (₫)"], 1):
        _c(ws, 5, col, title, bold=True, col=HDR_WHITE, fill=HDR_GREEN, sz=10, wrap=True)

    r_idx = 6
    if not rows:
        _mrow(ws, r_idx, 1, 5, "Không có dữ liệu trong kỳ.", italic=True)
        r_idx += 1
    else:
        for row in rows:
            ws.row_dimensions[int(r_idx)].height = 18
            is_thu = row["loai"] == "Thu"
            fill_bg = "f0faf5" if is_thu else "fff8f8"
            dg = f"{row['ma_phong']}"
            if row["ghi_chu"].strip():
                dg += f" · {row['ghi_chu']}"
            _c(ws, r_idx, 1, row["ngay"].strftime("%d/%m/%Y"), fill=fill_bg)
            _c(ws, r_idx, 2, row["loai"], col="1a5c3a" if is_thu else "b52b2b",
               bold=True, sz=9, fill=fill_bg)
            _c(ws, r_idx, 3, row["danh_muc"], ha="left", fill=fill_bg)
            _c(ws, r_idx, 4, dg, ha="left", fill=fill_bg)
            amt = row["so_tien"] if is_thu else -row["so_tien"]
            _c(ws, r_idx, 5, amt, ha="right", nf="#,##0;[Red]-#,##0",
               col="1a5c3a" if is_thu else "b52b2b", fill=fill_bg)
            r_idx += 1

    rn_final = int(r_idx)
    ws.row_dimensions[rn_final].height = 22
    _mrow(ws, rn_final, 1, 4, "Số dư cuối kỳ", bold=True, sz=11, ha="right",
          col="1a5c3a" if so_du >= 0 else "b52b2b", fill="d6f0e5")
    _c(ws, rn_final, 5, int(so_du), bold=True, sz=11, ha="right",
       col="1a5c3a" if so_du >= 0 else "b52b2b",
       nf="#,##0;[Red]-#,##0", fill="d6f0e5")

    ws.freeze_panes = "A6"
    buf = io.BytesIO(); wb.save(buf); return buf.getvalue()


def xuat_s1a(df_all, s, e) -> bytes:
    rows = []
    for _, row in df_all.iterrows():
        if (str(row.get("loai","")) == "Thu"
                and str(row.get("danh_muc","")) == "Doanh thu tiền phòng"
                and str(row.get("trang_thai","")) in TRANG_THAI_DA_THU
                and row["ngay"].date() >= s
                and row["ngay"].date() <= e):
            rows.append({
                "ngay": row["ngay"],
                "danh_muc": str(row["danh_muc"]),
                "ma_phong": str(row["ma_phong"]),
                "ghi_chu": str(row.get("ghi_chu", "") or ""),
                "so_tien": int(float(str(row["so_tien"]).replace(",","").replace(".","") or 0))
                    if not isinstance(row["so_tien"], (int, float))
                    else int(row["so_tien"]),
            })
    rows.sort(key=lambda x: x["ngay"])

    wb = Workbook(); ws = wb.active; ws.title = "S1a-HKD"
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 22
    _mrow(ws,1,1,3,"Mẫu số S1a-HKD", sz=9, ha="right", italic=True)
    _mrow(ws,2,1,3,"(Kèm theo Thông tư số 152/2025/TT-BTC ngày 31 tháng 12 năm 2025)", sz=9, ha="right", italic=True)
    ws.row_dimensions[3].height = 32
    _mrow(ws,3,1,3,"SỔ DOANH THU BÁN HÀNG HÓA, DỊCH VỤ", bold=True, sz=14, ha="center")
    _mrow(ws,4,1,3,f"Từ ngày {s.strftime('%d/%m/%Y')} đến ngày {e.strftime('%d/%m/%Y')}", sz=10, ha="center", italic=True)
    HDR = "1a7f5a"
    ws.row_dimensions[6].height = 32
    _c(ws,6,1,"Ngày tháng\n(Cột A)", bold=True, col="FFFFFF", fill=HDR, wrap=True)
    _c(ws,6,2,"Diễn giải\n(Cột B)",  bold=True, col="FFFFFF", fill=HDR, wrap=True)
    _c(ws,6,3,"Số tiền (₫)\n(Cột 1)",bold=True, col="FFFFFF", fill=HDR, wrap=True)
    tong = 0
    r_idx = 7
    if not rows:
        _mrow(ws,r_idx,1,3,"Không có dữ liệu trong kỳ.", italic=True)
        r_idx += 1
    else:
        for row in rows:
            ws.row_dimensions[r_idx].height = 18
            dg = f"{row['danh_muc']} — {row['ma_phong']}"
            if row["ghi_chu"].strip(): dg += f" ({row['ghi_chu']})"
            _c(ws, r_idx, 1, row["ngay"].strftime("%d/%m/%Y"))
            _c(ws, r_idx, 2, dg, ha="left")
            _c(ws, r_idx, 3, row["so_tien"], ha="right", nf="#,##0")
            tong += row["so_tien"]
            r_idx += 1
    ws.row_dimensions[r_idx].height = 24
    _c(ws, r_idx, 1, "", fill="d6f0e5")
    _c(ws, r_idx, 2, "Tổng cộng", bold=True, sz=11, ha="right", fill="d6f0e5")
    _c(ws, r_idx, 3, tong, bold=True, sz=11, col="1a5c3a", ha="right", nf="#,##0", fill="d6f0e5")
    ws.freeze_panes = "A7"
    buf = io.BytesIO(); wb.save(buf); return buf.getvalue()

def xuat_s2a(df_all, s, e) -> bytes:
    rows = []
    for _, row in df_all.iterrows():
        if (str(row.get("loai","")) == "Thu"
                and str(row.get("danh_muc","")) == "Doanh thu tiền phòng"
                and str(row.get("trang_thai","")) in TRANG_THAI_DA_THU
                and row["ngay"].date() >= s
                and row["ngay"].date() <= e):
            rows.append({
                "ngay": row["ngay"],
                "ma_phong": str(row["ma_phong"]),
                "ghi_chu": str(row.get("ghi_chu", "") or ""),
                "so_tien": int(row["so_tien"]) if isinstance(row["so_tien"], (int, float)) else 0,
            })
    rows.sort(key=lambda x: x["ngay"])

    wb = Workbook(); ws = wb.active; ws.title = "S2a-HKD"
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 50
    ws.column_dimensions["D"].width = 22
    _mrow(ws,1,1,4,"Mẫu số S2a-HKD",sz=9,ha="right",italic=True)
    _mrow(ws,2,1,4,"(Kèm theo Thông tư số 152/2025/TT-BTC ngày 31 tháng 12 năm 2025)",sz=9,ha="right",italic=True)
    ws.row_dimensions[3].height = 32
    _mrow(ws,3,1,4,"SỔ DOANH THU BÁN HÀNG HÓA, DỊCH VỤ",bold=True,sz=14,ha="center")
    _mrow(ws,4,1,4,f"Từ ngày {s.strftime('%d/%m/%Y')} đến ngày {e.strftime('%d/%m/%Y')}",sz=10,ha="center",italic=True)
    ws.row_dimensions[6].height = 22
    _mrow(ws,6,1,4,"1. Ngành nghề: Dịch vụ lưu trú khác",bold=True,sz=11,col="1a5c3a",ha="left",fill="d6f0e5")
    HDR = "1a7f5a"
    ws.row_dimensions[7].height = 32
    _c(ws,7,1,"STT",bold=True,col="FFFFFF",fill=HDR,wrap=True)
    _c(ws,7,2,"Ngày tháng",bold=True,col="FFFFFF",fill=HDR,wrap=True)
    _c(ws,7,3,"Diễn giải",bold=True,col="FFFFFF",fill=HDR,wrap=True)
    _c(ws,7,4,"Số tiền (₫)",bold=True,col="FFFFFF",fill=HDR,wrap=True)
    stt = 1; tong = 0; r_idx = 8
    if not rows:
        _mrow(ws,r_idx,1,4,"Không có dữ liệu doanh thu tiền phòng trong kỳ.",italic=True)
        r_idx += 1
    else:
        for row in rows:
            ws.row_dimensions[r_idx].height = 18
            dg = f"Tiền phòng {row['ma_phong']}"
            if row["ghi_chu"].strip(): dg += f" — {row['ghi_chu']}"
            _c(ws,r_idx,1,stt)
            _c(ws,r_idx,2,row["ngay"].strftime("%d/%m/%Y"))
            _c(ws,r_idx,3,dg,ha="left")
            _c(ws,r_idx,4,row["so_tien"],ha="right",nf="#,##0")
            tong += row["so_tien"]; stt += 1; r_idx += 1
    gtgt = int(round(tong * 0.05))
    tncn = int(round(tong * 0.02))
    def trow(tr, lbl, val, col="000000", fill="fff3cd", bold=True):
        ws.row_dimensions[tr].height = 22
        ws.merge_cells(start_row=tr,start_column=1,end_row=tr,end_column=3)
        c1 = ws.cell(row=tr,column=1,value=str(lbl))
        c1.font = Font(name="Arial",bold=bold,size=10,color=col)
        c1.alignment = Alignment(horizontal="right",vertical="center")
        c1.border = _bd(); c1.fill = PatternFill("solid",start_color=fill)
        for ci in [2,3]:
            cc = ws.cell(row=tr,column=ci); cc.border = _bd(); cc.fill = PatternFill("solid",start_color=fill)
        vc = ws.cell(row=tr,column=4,value=int(val))
        vc.font = Font(name="Arial",bold=bold,size=10,color=col)
        vc.alignment = Alignment(horizontal="right",vertical="center")
        vc.number_format = "#,##0"; vc.border = _bd(); vc.fill = PatternFill("solid",start_color=fill)
    trow(r_idx,   "Tổng cộng (1)",              tong, col="1a5c3a", fill="d6f0e5")
    trow(r_idx+1, "Thuế GTGT = (1) × 5%",       gtgt, col="c0392b", fill="fde8e8")
    trow(r_idx+2, "Thuế TNCN = (1) × 2%",       tncn, col="c0392b", fill="fde8e8")
    trow(r_idx+3, "Tổng số thuế GTGT phải nộp", gtgt, col="8b0000", fill="fcc8c8")
    trow(r_idx+4, "Tổng số thuế TNCN phải nộp", tncn, col="8b0000", fill="fcc8c8")
    ws.freeze_panes = "A8"
    buf = io.BytesIO(); wb.save(buf); return buf.getvalue()


# ══════════════════════════════════════════════════════════════════════════════
# PAGE CONFIG + CSS (UI v4.1 — Refined SaaS)
# ══════════════════════════════════════════════════════════════════════════════
st.set_page_config(page_title="Quản Lý Căn Hộ", page_icon="🏢",
                   layout="wide", initial_sidebar_state="expanded")

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Be+Vietnam+Pro:ital,wght@0,300;0,400;0,500;0,600;0,700;0,800;1,400&display=swap');

html, body, [class*="css"] {
    font-family: 'Be Vietnam Pro', sans-serif;
}

/* ── App Header ─────────────────────────────────────────────────────────── */
.app-hdr {
    background: linear-gradient(135deg, #1a7f5a 0%, #0e5c40 60%, #0a4530 100%);
    border-radius: 16px;
    padding: 22px 32px;
    margin-bottom: 20px;
    display: flex;
    align-items: center;
    gap: 18px;
    box-shadow: 0 4px 20px rgba(26, 127, 90, 0.25);
}
.app-hdr-icon {
    width: 52px; height: 52px;
    background: rgba(255,255,255,0.15);
    border-radius: 14px;
    display: flex; align-items: center; justify-content: center;
    font-size: 1.6rem; flex-shrink: 0;
    backdrop-filter: blur(4px);
}
.app-hdr h1 { color: white; margin: 0; font-size: 1.5rem; font-weight: 800; letter-spacing: -0.3px; }
.app-hdr p  { color: rgba(255,255,255,.65); margin: 3px 0 0; font-size: .8rem; font-weight: 400; }
.app-hdr-right { margin-left: auto; text-align: right; }
.app-hdr-date  { color: rgba(255,255,255,.5); font-size: .75rem; }
.app-hdr-version {
    display: inline-block; background: rgba(255,255,255,.15);
    border-radius: 20px; padding: 2px 10px;
    font-size: .72rem; color: rgba(255,255,255,.85); font-weight: 600;
    margin-top: 4px;
}

/* ── Connection Badge ───────────────────────────────────────────────────── */
.gs-badge {
    display: inline-flex; align-items: center; gap: 7px;
    background: #edfaf3; border: 1px solid #a8e6c8;
    border-radius: 20px; padding: 5px 14px;
    font-size: .78rem; font-weight: 600; color: #1a7f5a;
    margin-bottom: 18px;
    box-shadow: 0 1px 4px rgba(26,127,90,.08);
}

/* ── Tabs ───────────────────────────────────────────────────────────────── */
.stTabs [data-baseweb="tab-list"] {
    background: #f2f5f3; border-radius: 12px; padding: 4px; gap: 2px;
}
.stTabs [data-baseweb="tab"] {
    border-radius: 9px; font-weight: 600; font-size: .87rem; color: #777;
    padding: 8px 20px; transition: all .15s;
}
.stTabs [aria-selected="true"] {
    background: white !important; color: #1a7f5a !important;
    box-shadow: 0 2px 8px rgba(0,0,0,.1) !important;
}

/* ── Dashboard Cards (3-card layout) ───────────────────────────────────── */
.dash-card {
    border-radius: 16px; padding: 20px 22px; position: relative;
    overflow: hidden; transition: transform .18s, box-shadow .18s;
    cursor: default;
}
.dash-card:hover {
    transform: translateY(-2px);
    box-shadow: 0 8px 28px rgba(0,0,0,.12) !important;
}
.dash-card-thu {
    background: linear-gradient(135deg, #f0faf5 0%, #e2f5ec 100%);
    border: 1.5px solid #b8e6d0;
    box-shadow: 0 2px 10px rgba(26,158,95,.1);
}
.dash-card-chi {
    background: linear-gradient(135deg, #fff5f5 0%, #fee8e8 100%);
    border: 1.5px solid #f5c6c6;
    box-shadow: 0 2px 10px rgba(224,82,82,.1);
}
.dash-card-ton {
    background: linear-gradient(135deg, #1a7f5a 0%, #0e5c40 100%);
    border: 1.5px solid #167a56;
    box-shadow: 0 4px 16px rgba(26,127,90,.3);
}
.dash-card-label {
    font-size: .72rem; font-weight: 700; letter-spacing: .6px;
    text-transform: uppercase; margin-bottom: 8px;
}
.dash-card-thu  .dash-card-label { color: #1a7f5a; }
.dash-card-chi  .dash-card-label { color: #c0392b; }
.dash-card-ton  .dash-card-label { color: rgba(255,255,255,.7); }

.dash-card-amount {
    font-size: 1.55rem; font-weight: 800; letter-spacing: -1px; line-height: 1.1;
}
.dash-card-thu  .dash-card-amount { color: #1a7f5a; }
.dash-card-chi  .dash-card-amount { color: #c0392b; }
.dash-card-ton  .dash-card-amount { color: white; }

.dash-card-sub {
    font-size: .72rem; margin-top: 5px; font-weight: 500;
}
.dash-card-thu .dash-card-sub  { color: #5aab82; }
.dash-card-chi .dash-card-sub  { color: #e07878; }
.dash-card-ton .dash-card-sub  { color: rgba(255,255,255,.55); }

.dash-card-icon {
    position: absolute; right: 18px; top: 18px;
    font-size: 1.8rem; opacity: .18;
}

/* ── Pending Badge ──────────────────────────────────────────────────────── */
.pending-badge {
    display: inline-flex; align-items: center; gap: 8px;
    background: #fffbec; border: 1.5px solid #f5d167;
    border-radius: 10px; padding: 9px 16px; margin: 8px 0 16px;
    font-size: .84rem; color: #8a6000; font-weight: 600;
    width: 100%; box-sizing: border-box;
}
.pending-badge b { color: #6a4800; }

/* ── Balance Scoreboard (Tab 3) ─────────────────────────────────────────── */
.sb  { display: flex; gap: 10px; margin: 12px 0; flex-wrap: wrap; }
.sc  { flex: 1; min-width: 130px; border-radius: 13px; padding: 13px 16px; border-left: 4px solid; }
.sc.dau { background: #f7f8f7; border-color: #c8ccc9; }
.sc.thu { background: #e8f5f0; border-color: #1a9e5f; }
.sc.chi { background: #fdf0f0; border-color: #e05252; }
.sc.end { background: linear-gradient(135deg,#1a7f5a,#0f5c40); border-color: #1a7f5a; }
.sc .sl { font-size: .68rem; font-weight: 700; letter-spacing: .5px; text-transform: uppercase; }
.sc.dau .sl, .sc.thu .sl, .sc.chi .sl { color: #999; }
.sc.end .sl { color: rgba(255,255,255,.65); }
.sc .sv { font-size: 1.15rem; font-weight: 800; margin-top: 4px; }
.sc.dau .sv { color: #555; }
.sc.thu .sv { color: #1a9e5f; }
.sc.chi .sv { color: #e05252; }
.sc.end .sv { color: white; }

/* ── Transaction Row (tx) ───────────────────────────────────────────────── */
.tx {
    background: white; border-radius: 11px; padding: 11px 14px; margin-bottom: 6px;
    display: flex; align-items: center; gap: 10px;
    box-shadow: 0 1px 5px rgba(0,0,0,.06); border: 1px solid #f0f0f0;
    transition: box-shadow .15s;
}
.tx:hover { box-shadow: 0 3px 12px rgba(0,0,0,.1); }
.ti { width: 36px; height: 36px; border-radius: 50%; display: flex; align-items: center;
      justify-content: center; font-size: .95rem; flex-shrink: 0; }
.ti.thu { background: #e8f5f0; } .ti.chi { background: #fdf0f0; }
.tn { flex: 1; min-width: 0; }
.tc { font-size: .84rem; font-weight: 600; color: #222; }
.tm { font-size: .73rem; color: #aaa; margin-top: 1px; }
.ta { font-size: .95rem; font-weight: 700; white-space: nowrap; }
.ta.thu { color: #1a9e5f; } .ta.chi { color: #e05252; }

/* ── Mini Stats (side panel) ────────────────────────────────────────────── */
.ms { border-radius: 12px; padding: 12px 16px; margin-bottom: 8px; border-left: 4px solid; }
.ms.thu { background: #e8f5f0; border-color: #1a9e5f; }
.ms.chi { background: #fdf0f0; border-color: #e05252; }
.ms .lbl { font-size: .68rem; font-weight: 700; text-transform: uppercase; letter-spacing: .5px; }
.ms.thu .lbl { color: #1a7f5a; } .ms.chi .lbl { color: #e05252; }
.ms .val { font-size: 1.1rem; font-weight: 800; margin-top: 2px; }
.ms.thu .val { color: #1a9e5f; } .ms.chi .val { color: #e05252; }

/* ── Section Heading ────────────────────────────────────────────────────── */
.section-heading {
    font-size: 1rem; font-weight: 700; color: #1a2e24;
    margin: 2px 0 14px; padding-bottom: 8px;
    border-bottom: 2px solid #e8f0ec;
    display: flex; align-items: center; gap: 8px;
}
.section-heading span { font-size: 1.05rem; }

/* ── Form compact blocks ────────────────────────────────────────────────── */
.form-group-label {
    font-size: .72rem; font-weight: 700; color: #888;
    letter-spacing: .5px; text-transform: uppercase;
    margin: 14px 0 6px;
}
.chi-notice {
    background: #fff5f5; border-left: 3px solid #e05252;
    border-radius: 10px; padding: 10px 14px; margin: 8px 0 14px;
    font-size: .84rem; color: #c0392b; font-weight: 500;
}

/* ── Import Box ─────────────────────────────────────────────────────────── */
.import-box {
    background: linear-gradient(135deg, #f0f7ff, #e8f3ff);
    border: 2px dashed #5a9fff; border-radius: 14px;
    padding: 18px 22px; margin: 12px 0 18px;
}
.import-box b { color: #2563c0; }

/* ── Pending Box ────────────────────────────────────────────────────────── */
.cho-tt-box {
    background: linear-gradient(135deg, #fffbec, #fff8e0);
    border: 1.5px solid #f5c842; border-radius: 14px;
    padding: 16px 20px; margin: 12px 0 18px;
}
.cho-tt-title { font-size: .95rem; font-weight: 700; color: #8a6000; margin-bottom: 8px; }

/* ── Edit Form ──────────────────────────────────────────────────────────── */
.edit-form {
    background: #f6fffc; border: 2px solid #1a9e5f;
    border-radius: 14px; padding: 18px 20px; margin: 8px 0 16px;
}
.edit-form-title { font-size: .92rem; font-weight: 700; color: #1a7f5a; margin-bottom: 12px; }

/* ── Tax Info ───────────────────────────────────────────────────────────── */
.tax-info {
    background: linear-gradient(135deg, #fff8e6, #fff4d6);
    border-left: 4px solid #f5a623; border-radius: 10px;
    padding: 12px 16px; margin-bottom: 16px; font-size: .87rem;
}

/* ── Tiền Cọc ───────────────────────────────────────────────────────────── */
.coc-card {
    background: linear-gradient(135deg, #f4f0ff 0%, #ede8ff 100%);
    border: 1.5px solid #c9baff; border-radius: 14px;
    padding: 16px 20px; margin-bottom: 14px;
    box-shadow: 0 2px 10px rgba(120,80,220,.09);
}
.coc-header {
    font-size: .72rem; font-weight: 700; letter-spacing: .5px;
    text-transform: uppercase; color: #7c5cbf; margin-bottom: 6px;
}
.coc-room { font-size: 1.1rem; font-weight: 800; color: #3d1f8e; }
.coc-amount { font-size: 1.3rem; font-weight: 800; color: #5b2dd9; margin-top: 2px; }
.coc-meta { font-size: .75rem; color: #9e8ec5; margin-top: 4px; }
.coc-warn { font-size: .82rem; color: #b45309; font-weight: 600; margin-top: 4px; }
.coc-empty { opacity: .45; }
.coc-total-box {
    background: linear-gradient(135deg, #5b2dd9 0%, #3d1f8e 100%);
    border-radius: 14px; padding: 18px 24px; margin: 6px 0 20px;
    color: white; box-shadow: 0 4px 18px rgba(91,45,217,.28);
}
.coc-total-label { font-size: .72rem; font-weight: 700; letter-spacing: .6px;
    text-transform: uppercase; color: rgba(255,255,255,.65); margin-bottom: 6px; }
.coc-total-amount { font-size: 1.9rem; font-weight: 800; letter-spacing: -1px; }
.coc-total-sub { font-size: .75rem; color: rgba(255,255,255,.5); margin-top: 4px; }
.coc-form-box {
    background: #faf8ff; border: 2px solid #c9baff;
    border-radius: 14px; padding: 20px 22px;
}
.coc-form-title { font-size: .95rem; font-weight: 700; color: #5b2dd9; margin-bottom: 14px; }

/* ── Cọc Grid — CSS-only, đảm bảo thứ tự row-first trên mọi màn hình ─── */
.coc-grid {
    display: grid;
    grid-template-columns: repeat(3, 1fr);
    gap: 14px;
    margin-bottom: 20px;
}
@media (max-width: 768px) {
    .coc-grid { grid-template-columns: 1fr; }
}

/* ── Stagger column for tx actions ─────────────────────────────────────── */
div[data-testid="stHorizontalBlock"]:has(.tx) {
    align-items: center !important; gap: 6px !important;
}
div[data-testid="stHorizontalBlock"]:has(.tx) > div[data-testid="column"]:last-child {
    flex: 0 0 32px !important; min-width: 32px !important;
    max-width: 32px !important; padding: 0 !important;
}
div[data-testid="stHorizontalBlock"]:has(.tx) > div[data-testid="column"]:last-child button {
    background: #f0f0f0 !important; border: 1px solid #ddd !important;
    border-radius: 8px !important; color: #888 !important;
    font-size: .75rem !important; min-height: 44px !important;
    height: 44px !important; width: 32px !important;
    padding: 0 !important; box-shadow: none !important; transform: none !important;
}
div[data-testid="stHorizontalBlock"]:has(.tx) > div[data-testid="column"]:last-child button:hover {
    background: #e8f5ee !important; border-color: #1a9e5f !important;
    color: #1a7f5a !important; transform: none !important;
}

/* ── Global Buttons ─────────────────────────────────────────────────────── */
.stButton > button {
    background: #1a7f5a; color: white; border: none; border-radius: 10px;
    font-family: 'Be Vietnam Pro', sans-serif; font-weight: 600;
    font-size: .9rem; padding: 8px 22px; width: 100%;
    transition: background .15s, transform .15s, box-shadow .15s;
}
.stButton > button:hover {
    background: #15694a; transform: translateY(-1px);
    box-shadow: 0 4px 12px rgba(26,127,90,.25);
}

/* ── Scrollbar ──────────────────────────────────────────────────────────── */
::-webkit-scrollbar { width: 5px; }
::-webkit-scrollbar-thumb { background: #c4e0d6; border-radius: 3px; }

#MainMenu, footer, header { visibility: hidden; }
</style>
""", unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════════════════════
# INIT SESSION STATE
# ══════════════════════════════════════════════════════════════════════════════
_defaults = {
    "room":           ROOMS[0],
    "excel_preview":  None,
    "excel_thang":    None,
    "ky2":            "Tháng này",
    "nam2":           date.today().year,
    "edit_id":        None,
    "form_ver":       0,
    "coc_form_ver":   0,
    "coc_live_tien":  "",
    "nav_year":       date.today().year,
    "nav_month":      date.today().month,
    "live_tien_thu":  "",
    "live_tien_chi":  "",
    "selected_gd":    None,
}
for k, v in _defaults.items():
    if k not in st.session_state:
        st.session_state[k] = v

today = date.today()

# ── HEADER ───────────────────────────────────────────────────────────────────
st.markdown(f"""
<div class="app-hdr">
  <div class="app-hdr-icon">🏢</div>
  <div>
    <h1>Quản Lý Thu Chi Căn Hộ</h1>
    <p>Google Sheets · 9 phòng · TT152/2025/TT-BTC</p>
  </div>
  <div class="app-hdr-right">
    <div class="app-hdr-date">{today.strftime('%d/%m/%Y')}</div>
    <div class="app-hdr-version">v4.2</div>
  </div>
</div>
""", unsafe_allow_html=True)

ws_check = get_worksheet()
if ws_check:
    st.markdown('<div class="gs-badge">🟢 Google Sheets đã kết nối — Dữ liệu được lưu vĩnh viễn</div>', unsafe_allow_html=True)
else:
    st.warning("⚠️ Chưa kết nối Google Sheets. Hãy cấu hình `secrets.toml`.")

st.sidebar.markdown("---")
st.sidebar.subheader("🔍 Test Cấu Hình Phòng")
phong_test = st.sidebar.selectbox("Chọn phòng để test data:", ROOMS, key="sb_phong_test")
# Gọi hàm get_thong_tin_phong để lấy dữ liệu thực tế từ Google Sheets
data_phong = get_thong_tin_phong(phong_test)
st.sidebar.json(data_phong)

# ══════════════════════════════════════════════════════════════════════════════
# TABS
# ══════════════════════════════════════════════════════════════════════════════
tab1, tab2, tab3, tab4, tab5 = st.tabs([
    "➕  Nhập Liệu",
    "📤  Import & Thu Tiền",
    "📒  Sổ Dòng Tiền",
    "📊  Xuất Sổ Thuế",
    "📌  Quản lý Tiền Cọc",
])

# ════════════════════════════════════════════════════════════════════════════
# TAB 1 — NHẬP LIỆU NHANH
# ════════════════════════════════════════════════════════════════════════════
with tab1:
    @st.fragment
    def fragment_nhap_lieu():
        df_all = load_all()
        col_main, col_side = st.columns([5, 3], gap="large")

        with col_main:
            st.markdown('<div class="section-heading"><span>➕</span> Thêm giao dịch mới</div>', unsafe_allow_html=True)

            # ── Row 1: Loại + Ngày ───────────────────────────────────────────
            r1a, r1b = st.columns([1, 2])
            with r1a:
                loai_sel = st.radio("Loại giao dịch", ["Thu", "Chi"], horizontal=True, key="loai_r")
            with r1b:
                ngay_sel = st.date_input("📅 Ngày giao dịch", value=today, key="ngay_i")

            if loai_sel == "Thu":
                # ── Row 2: Chọn phòng ─────────────────────────────────────────
                st.markdown('<div class="form-group-label">Chọn nhanh phòng</div>', unsafe_allow_html=True)
                room_cols = st.columns(len(ROOMS))
                for i, p in enumerate(ROOMS):
                    with room_cols[i]:
                        if st.button(p, key=f"rb_{p}", use_container_width=True,
                                     type="primary" if st.session_state.room == p else "secondary"):
                            st.session_state.room = p
                            st.rerun(scope="fragment")

                ma_phong_sel = st.session_state.room
                st.markdown(f"<div style='font-size:.8rem;color:#1a7f5a;font-weight:600;margin:4px 0 10px'>📍 Đang chọn: <b>{ma_phong_sel}</b></div>", unsafe_allow_html=True)

                fv = st.session_state.form_ver

                # ── Row 3: Danh mục + Số tiền ─────────────────────────────────
                r3a, r3b = st.columns([2, 3])
                with r3b:
                    so_tien_raw = st.text_input(
                        "💰 Số tiền (₫)",
                        placeholder="VD: 3.500.000",
                        key=f"st_thu_{fv}",
                        on_change=lambda: st.session_state.update(
                            live_tien_thu=st.session_state.get(f"st_thu_{fv}", "")
                        ),
                    )
                    so_tien = parse_tien(so_tien_raw)
                    if so_tien > 0:
                        st.markdown(f"<div style='color:#1a9e5f;font-weight:800;font-size:1.25rem;margin:-2px 0 4px'>{fmt_vnd(so_tien)}</div>", unsafe_allow_html=True)

                # ── Row 4: Ghi chú + Submit (trong form) ──────────────────────
                with st.form(key=f"form_thu_{fv}", clear_on_submit=True):
                    with r3a:
                        danh_muc_sel = st.selectbox("📂 Danh mục", DANH_MUC_THU, key="dm_thu")
                    ghi_chu = st.text_input("📝 Ghi chú (tùy chọn)", placeholder="Tháng ..., nội dung ...", key=f"gc_thu_{fv}")
                    submitted_thu = st.form_submit_button("💾  Lưu giao dịch Thu", type="primary", use_container_width=True)

                if submitted_thu:
                    so_tien_submit = parse_tien(st.session_state.get(f"st_thu_{fv}", ""))
                    if so_tien_submit <= 0:
                        st.error("⚠️ Nhập số tiền lớn hơn 0")
                    else:
                        ok = insert_gd(ngay_sel, "Thu", ma_phong_sel, danh_muc_sel, so_tien_submit, ghi_chu, trang_thai="Đã thu")
                        if ok:
                            st.toast(f"✅ Đã lưu: {danh_muc_sel} · {ma_phong_sel} · {fmt_vnd(so_tien_submit)}", icon="💚")
                            st.session_state.live_tien_thu = ""
                            st.session_state.form_ver += 1
                            st.rerun(scope="fragment")

            else:
                # ── Chi nhanh ─────────────────────────────────────────────────
                st.markdown('<div class="chi-notice">⚡ Chế độ Chi nhanh — Nhập <b>Số tiền</b> &amp; <b>Ghi chú</b> rồi Lưu!</div>', unsafe_allow_html=True)

                fv = st.session_state.form_ver

                # ── Row 2: Danh mục + Số tiền ─────────────────────────────────
                r2a, r2b = st.columns([2, 3])
                with r2b:
                    so_tien_chi_raw = st.text_input(
                        "💰 Số tiền (₫)",
                        placeholder="VD: 700.000",
                        key=f"st_chi_{fv}",
                        on_change=lambda: st.session_state.update(
                            live_tien_chi=st.session_state.get(f"st_chi_{fv}", "")
                        ),
                    )
                    so_tien_chi = parse_tien(so_tien_chi_raw)
                    if so_tien_chi > 0:
                        st.markdown(f"<div style='color:#e05252;font-weight:800;font-size:1.25rem;margin:-2px 0 4px'>{fmt_vnd(so_tien_chi)}</div>", unsafe_allow_html=True)

                with st.form(key=f"form_chi_{fv}", clear_on_submit=True):
                    with r2a:
                        dm_chi = st.selectbox("📂 Danh mục chi", DANH_MUC_CHI, index=0, key="dm_chi")
                    gc_chi = st.text_input("📝 Ghi chú — Mua / chi gì?", key=f"gc_chi_{fv}")
                    submitted_chi = st.form_submit_button("💾  Lưu khoản Chi", type="primary", use_container_width=True)

                if submitted_chi:
                    so_tien_chi_submit = parse_tien(st.session_state.get(f"st_chi_{fv}", ""))
                    if so_tien_chi_submit <= 0:
                        st.error("⚠️ Nhập số tiền lớn hơn 0")
                    else:
                        ok = insert_gd(ngay_sel, "Chi", "Chung", dm_chi, so_tien_chi_submit, gc_chi, trang_thai="Đã thu")
                        if ok:
                            st.toast(f"✅ Đã lưu: {dm_chi} · {fmt_vnd(so_tien_chi_submit)}", icon="🔴")
                            st.session_state.live_tien_chi = ""
                            st.session_state.form_ver += 1
                            st.rerun(scope="fragment")

        # ── Side Panel: Dashboard Cards ───────────────────────────────────────
        with col_side:
            sd_dau, t_thu, t_chi, sd_cuoi = compute_balances(
                df_all,
                date(today.year, today.month, 1),
                today
            )

            # Pending badge
            df_cho_side = df_all[df_all["trang_thai"] == "Chờ thanh toán"] if not df_all.empty else pd.DataFrame()
            if not df_cho_side.empty:
                tong_cho_side = df_cho_side["so_tien"].sum()
                st.markdown(f"""
                <div class="pending-badge">
                    ⏳ Chờ thanh toán: <b>{fmt_vnd(tong_cho_side)}</b>
                    &nbsp;·&nbsp; {len(df_cho_side)} GD
                </div>
                """, unsafe_allow_html=True)

            # 3 Dashboard Cards
            st.markdown(f"""
            <div class="dash-card dash-card-thu" style="margin-bottom:10px">
                <div class="dash-card-icon">📈</div>
                <div class="dash-card-label">Tổng Thu Thực Tế</div>
                <div class="dash-card-amount">+{fmt_vnd(t_thu)}</div>
                <div class="dash-card-sub">Tháng {today.month:02d}/{today.year}</div>
            </div>
            <div class="dash-card dash-card-chi" style="margin-bottom:10px">
                <div class="dash-card-icon">📉</div>
                <div class="dash-card-label">Tổng Chi Thực Tế</div>
                <div class="dash-card-amount">-{fmt_vnd(t_chi)}</div>
                <div class="dash-card-sub">Tháng {today.month:02d}/{today.year}</div>
            </div>
            <div class="dash-card dash-card-ton" style="margin-bottom:14px">
                <div class="dash-card-icon" style="opacity:.2">🏦</div>
                <div class="dash-card-label">Tồn Quỹ Hiện Tại</div>
                <div class="dash-card-amount">{fmt_vnd(sd_cuoi)}</div>
                <div class="dash-card-sub">Số dư đầu kỳ: {fmt_vnd(sd_dau)}</div>
            </div>
            """, unsafe_allow_html=True)

            # Recent transactions
            st.markdown("<div style='font-size:.72rem;font-weight:700;color:#bbb;letter-spacing:.5px;margin-bottom:8px'>GẦN ĐÂY (đã thu)</div>", unsafe_allow_html=True)
            df_recent = df_all[df_all["trang_thai"].isin(TRANG_THAI_DA_THU)] if not df_all.empty else pd.DataFrame()
            if not df_recent.empty:
                for _, r in df_recent.head(6).iterrows():
                    icon = "📈" if r["loai"] == "Thu" else "📉"
                    cls  = "thu" if r["loai"] == "Thu" else "chi"
                    pref = "+" if r["loai"] == "Thu" else "-"
                    note = f" · {r['ghi_chu']}" if str(r.get("ghi_chu", "")).strip() else ""
                    st.markdown(f"""
                    <div class="tx">
                        <div class="ti {cls}">{icon}</div>
                        <div class="tn">
                            <div class="tc">{r['danh_muc']}</div>
                            <div class="tm">{r['ma_phong']}{note} · {r['ngay'].strftime('%d/%m')}</div>
                        </div>
                        <div class="ta {cls}">{pref}{fmt_vnd(r['so_tien'])}</div>
                    </div>
                    """, unsafe_allow_html=True)
            else:
                st.info("Chưa có giao dịch thực thu.")

    fragment_nhap_lieu()


# ════════════════════════════════════════════════════════════════════════════
# TAB 2 — IMPORT EXCEL + CHỐT SỐ + XÁC NHẬN THU TIỀN
# ════════════════════════════════════════════════════════════════════════════
with tab2:

    # ── Fragment 2: Import Excel ──────────────────────────────────────────────
    @st.fragment
    def fragment_import_excel():
        st.markdown('<div class="section-heading"><span>📤</span> Import báo cáo tháng từ file Excel</div>', unsafe_allow_html=True)
        st.markdown("""
        <div class="import-box">
            <b>📋 Hướng dẫn:</b><br>
            1. Upload file Excel báo cáo tháng (sheet <b>'3. QL hóa đơn'</b>)<br>
            2. App tự đọc header dòng 2, bóc tách 9 phòng<br>
            3. Kiểm tra bảng xem trước → Bấm <b>"Xác nhận & Chốt số"</b><br>
            ⚠️ Dữ liệu sẽ lưu trạng thái <b>Chờ thanh toán</b> — chưa tính vào dòng tiền thực tế
        </div>
        """, unsafe_allow_html=True)

        ic1, ic2, ic3 = st.columns([3, 1, 1])
        with ic1:
            uploaded = st.file_uploader("📂 Chọn file Excel (.xlsx)", type=["xlsx"], key="xl_upload")
        with ic2:
            thang_imp = st.selectbox("Tháng", list(range(1, 13)),
                                     index=today.month - 1,
                                     format_func=lambda x: f"T{x:02d}", key="th_imp")
        with ic3:
            nam_imp = st.selectbox("Năm", list(range(2023, today.year + 2)),
                                   index=list(range(2023, today.year + 2)).index(today.year), key="nm_imp")

        if uploaded:
            with st.spinner("Đang đọc và phân tích file Excel..."):
                df_preview = parse_excel_invoice(uploaded, thang_imp, nam_imp)

            if df_preview is not None:
                st.session_state.excel_preview = df_preview
                st.session_state.excel_thang   = (thang_imp, nam_imp)

                st.markdown(f'<div class="section-heading"><span>🔍</span> Xem trước — Tháng {thang_imp:02d}/{nam_imp}</div>', unsafe_allow_html=True)
                st.markdown(f"Tìm thấy **{len(df_preview)} phòng**. Kiểm tra số liệu trước khi chốt:")

                show_df = df_preview.copy()
                rename_map = {
                    "ma_phong": "Phòng", "ten_kh": "Tên KH/ĐD",
                    "tien_phong": "Tiền nhà (₫)", "so_dien_kwh": "kWh",
                    "tien_dien": "Tiền điện (₫)", "tien_nuoc": "Tiền nước (₫)",
                    "tong_phai_thu": "Tổng phải thu (₫)", "thang_nhap": "Tháng",
                    "trang_thai": "Trạng thái",
                }
                show_df = show_df.rename(columns={k: v for k, v in rename_map.items() if k in show_df.columns})

                # Build column_config for st.dataframe
                col_cfg = {}
                for money_col in ["Tiền nhà (₫)", "Tiền điện (₫)", "Tiền nước (₫)", "Tổng phải thu (₫)"]:
                    if money_col in show_df.columns:
                        col_cfg[money_col] = st.column_config.NumberColumn(
                            money_col, format="%,.0f ₫", min_value=0
                        )
                if "Trạng thái" in show_df.columns:
                    col_cfg["Trạng thái"] = st.column_config.SelectboxColumn(
                        "Trạng thái",
                        options=["Chờ thanh toán", "Đã thanh toán", "Đã thu"],
                        disabled=True,
                    )

                st.dataframe(show_df, use_container_width=True, hide_index=True, column_config=col_cfg)

                if "tong_phai_thu" in df_preview.columns:
                    tong_kt = df_preview["tong_phai_thu"].sum()
                    st.markdown(f"""
                    <div style="background:#e8f5f0;border-radius:12px;padding:13px 18px;margin:10px 0;
                                display:flex;gap:24px;align-items:center;border:1px solid #b8e6d0;">
                        <div style="font-size:.88rem;color:#1a7f5a;font-weight:700">
                            ✅ Tổng phải thu {len(df_preview)} phòng:
                            <span style="font-size:1.2rem;font-weight:800"> {fmt_vnd(tong_kt)}</span>
                        </div>
                        <div style="font-size:.78rem;color:#8a6000;font-weight:500">
                            ⚠️ Trạng thái: <b>Chờ thanh toán</b>
                        </div>
                    </div>
                    """, unsafe_allow_html=True)

                st.markdown("---")
                col_btn1, col_btn2 = st.columns([2, 1])
                with col_btn1:
                    if st.button("🚀  Xác nhận & Chốt số vào Google Sheets", key="push_gs", type="primary"):
                        thang_m, nam_m = st.session_state.excel_thang
                        ngay_base = date(nam_m, thang_m, 1)
                        rows_to_push = []
                        df_p = st.session_state.excel_preview
                        thang_str = f"{thang_m:02d}/{nam_m}"

                        for _, row in df_p.iterrows():
                            ma_p = row.get("ma_phong", "Chung")
                            if row.get("tien_phong", 0) > 0:
                                rows_to_push.append([
                                    str(ngay_base), "Thu", ma_p, "Doanh thu tiền phòng",
                                    float(row["tien_phong"]), f"Import T{thang_m:02d}/{nam_m}",
                                    thang_str, "Chờ thanh toán"
                                ])
                            if row.get("tien_dien", 0) > 0:
                                kwh = row.get("so_dien_kwh", 0)
                                rows_to_push.append([
                                    str(ngay_base), "Thu", ma_p, "Thu hộ tiền điện",
                                    float(row["tien_dien"]),
                                    f"Điện {int(kwh)} kWh — Import T{thang_m:02d}/{nam_m}",
                                    thang_str, "Chờ thanh toán"
                                ])
                            if row.get("tien_nuoc", 0) > 0:
                                rows_to_push.append([
                                    str(ngay_base), "Thu", ma_p, "Thu hộ tiền nước",
                                    float(row["tien_nuoc"]),
                                    f"Nước — Import T{thang_m:02d}/{nam_m}",
                                    thang_str, "Chờ thanh toán"
                                ])

                        with st.spinner(f"Đang ghi {len(rows_to_push)} dòng vào Google Sheets..."):
                            ok = append_batch(rows_to_push)

                        if ok:
                            st.toast(f"✅ Đã chốt {len(rows_to_push)} GD tháng {thang_m:02d}/{nam_m}", icon="📊")
                            st.success(f"Chốt số thành công · Trạng thái: **Chờ thanh toán**\n\n👇 Kéo xuống mục **Xác nhận Thu Tiền** để tích phòng đã đóng tiền.")
                            st.balloons()
                            st.session_state.excel_preview = None
                            st.rerun(scope="fragment")
                        else:
                            st.error("❌ Lỗi khi ghi vào Google Sheets.")

                with col_btn2:
                    if st.button("🗑️  Huỷ, upload lại", key="cancel_imp"):
                        st.session_state.excel_preview = None
                        st.rerun(scope="fragment")

    fragment_import_excel()

    st.markdown("---")

    # ── Fragment 3: Xác nhận thu tiền ────────────────────────────────────────
    @st.fragment
    def fragment_xac_nhan_thu_tien():
        st.markdown('<div class="section-heading"><span>💰</span> Xác Nhận Thu Tiền Thực Tế</div>', unsafe_allow_html=True)
        st.markdown("""
        <div class="cho-tt-box">
            <div class="cho-tt-title">📋 Danh sách phòng đang Chờ thanh toán</div>
            Tích chọn những phòng đã thực sự đóng tiền → Bấm <b>"Xác nhận đã thu tiền"</b>
            để chuyển sang <b>Đã thanh toán</b> và ghi nhận vào dòng tiền thực tế.
        </div>
        """, unsafe_allow_html=True)

        df_all = load_all()

        df_cho = df_all[
            (df_all["trang_thai"] == "Chờ thanh toán") &
            (df_all["loai"] == "Thu")
        ].copy()

        if df_cho.empty:
            st.info("✅ Không có phòng nào đang chờ thanh toán.")
            return

        thangs = df_cho["thang_nhap"].dropna().unique().tolist()
        thangs.sort(reverse=True)

        thang_sel = st.selectbox(
            "🗓️ Chọn tháng cần xác nhận",
            thangs,
            format_func=lambda x: f"Tháng {x}",
            key="thang_sel_tt"
        )

        df_thang = df_cho[df_cho["thang_nhap"] == thang_sel].copy()

        summary = (
            df_thang.groupby("ma_phong")
            .agg(
                tong_phai_thu=("so_tien", "sum"),
                so_hang=("id", "count"),
                ids=("id", lambda x: list(x.astype(int))),
            )
            .reset_index()
        )

        st.markdown(f"**{len(summary)} phòng** đang chờ thanh toán tháng **{thang_sel}** — Tổng: **{fmt_vnd(summary['tong_phai_thu'].sum())}**")

        with st.expander("🗑️ Xóa dữ liệu tháng này (import nhầm?)", expanded=False):
            st.warning(f"⚠️ Sẽ xóa toàn bộ **{len(df_thang)} dòng** — tháng **{thang_sel}** · Không thể hoàn tác!")
            all_ids_thang = df_thang["id"].astype(int).tolist()
            if st.button(f"🗑️ Xóa toàn bộ {len(df_thang)} dòng tháng {thang_sel}", key="del_thang_cho"):
                with st.spinner("Đang xóa..."):
                    ok = delete_batch(all_ids_thang)
                if ok:
                    st.toast(f"🗑️ Đã xóa {len(all_ids_thang)} dòng tháng {thang_sel}", icon="🗑️")
                    st.rerun(scope="fragment")
                else:
                    st.error("❌ Lỗi xóa. Kiểm tra kết nối.")

        st.markdown("---")

        selected_ids = []
        cols_check = st.columns(3)
        for i, row in summary.iterrows():
            with cols_check[i % 3]:
                checked = st.checkbox(
                    f"**{row['ma_phong']}** — {fmt_vnd(row['tong_phai_thu'])}",
                    key=f"chk_{row['ma_phong']}_{thang_sel}",
                    help=f"{row['so_hang']} dòng giao dịch"
                )
                if checked:
                    selected_ids.extend(row["ids"])

        if selected_ids:
            n_phong_sel = len([r for _, r in summary.iterrows() if any(gid in selected_ids for gid in r["ids"])])
            tong_sel = summary[summary["ids"].apply(lambda ids: any(gid in selected_ids for gid in ids))]["tong_phai_thu"].sum()
            st.markdown(f"""
            <div style="background:#e8f5f0;border-radius:11px;padding:12px 18px;margin:10px 0;
                        font-size:.9rem;color:#1a7f5a;font-weight:600;border:1px solid #b8e6d0;">
                ✅ Đã chọn <b>{n_phong_sel} phòng</b> · Tổng thu: <b>{fmt_vnd(tong_sel)}</b>
            </div>
            """, unsafe_allow_html=True)

            if st.button("✅  Xác nhận đã thu tiền — Ghi nhận dòng tiền thực tế",
                         key="btn_xn_thu", type="primary"):
                with st.spinner("Đang cập nhật trạng thái lên Google Sheets..."):
                    ok = update_trang_thai(selected_ids, "Đã thanh toán")
                if ok:
                    st.toast(f"🎉 Xác nhận {n_phong_sel} phòng · {fmt_vnd(tong_sel)}", icon="✅")
                    st.rerun(scope="fragment")
                else:
                    st.error("❌ Lỗi cập nhật. Kiểm tra kết nối Google Sheets.")
        else:
            st.caption("👆 Tích chọn ít nhất 1 phòng để xác nhận.")

        df_da_tt = df_all[
            (df_all["trang_thai"] == "Đã thanh toán") &
            (df_all["loai"] == "Thu")
        ]
        if not df_da_tt.empty:
            with st.expander(f"📋 Xem giao dịch đã thanh toán ({len(df_da_tt)} dòng)", expanded=False):
                show = df_da_tt[["ngay", "ma_phong", "danh_muc", "so_tien", "thang_nhap"]].copy()
                col_cfg_da = {
                    "ngay":      st.column_config.DateColumn("Ngày", format="DD/MM/YYYY"),
                    "so_tien":   st.column_config.NumberColumn("Số tiền (₫)", format="%,.0f ₫"),
                    "ma_phong":  st.column_config.TextColumn("Phòng"),
                    "danh_muc":  st.column_config.TextColumn("Danh mục"),
                    "thang_nhap": st.column_config.TextColumn("Tháng"),
                }
                st.dataframe(show, use_container_width=True, hide_index=True, column_config=col_cfg_da)

    fragment_xac_nhan_thu_tien()


# ════════════════════════════════════════════════════════════════════════════
# TAB 3 — SỔ DÒNG TIỀN
# ════════════════════════════════════════════════════════════════════════════
with tab3:
    df_all2 = load_all()
    df_real2 = df_all2[df_all2["trang_thai"].isin(TRANG_THAI_DA_THU)].copy() if not df_all2.empty else df_all2.copy()

    st.markdown('<div class="section-heading"><span>📒</span> Sổ Dòng Tiền Thực Tế</div>', unsafe_allow_html=True)

    ff1, ff2 = st.columns(2)
    with ff1: phong2 = st.selectbox("🏠 Phòng", ["Tất cả"] + ROOMS, key="ph2")
    with ff2: loai2  = st.selectbox("Loại", ["Tất cả", "Thu", "Chi"], key="lo2")

    nav_y = st.session_state.nav_year
    nav_m = st.session_state.nav_month

    NAV_MONTHS = ["Tháng 1","Tháng 2","Tháng 3","Tháng 4","Tháng 5","Tháng 6",
                  "Tháng 7","Tháng 8","Tháng 9","Tháng 10","Tháng 11","Tháng 12"]
    PERIOD_PRESETS = ["Tháng này","Quý 1","Quý 2","Quý 3","Quý 4",
                      "6 Tháng đầu năm","6 Tháng cuối năm","Cả năm","Tùy chỉnh khoảng ngày"]

    nav_col1, nav_col2, nav_col3 = st.columns([1, 4, 1])
    with nav_col1:
        if st.button("◀", key="nav_prev", use_container_width=True):
            if nav_m == 1:
                st.session_state.nav_month = 12
                st.session_state.nav_year  = nav_y - 1
            else:
                st.session_state.nav_month = nav_m - 1
            st.rerun()
    with nav_col2:
        st.markdown(f"""
        <div style="text-align:center;font-size:1.08rem;font-weight:800;
                    color:#1a7f5a;padding:8px 0;line-height:1.2">
            {NAV_MONTHS[nav_m-1]}
            <span style="font-size:.88rem;font-weight:500;color:#777;margin-left:6px">{nav_y}</span>
        </div>
        """, unsafe_allow_html=True)
    with nav_col3:
        if st.button("▶", key="nav_next", use_container_width=True):
            if nav_m == 12:
                st.session_state.nav_month = 1
                st.session_state.nav_year  = nav_y + 1
            else:
                st.session_state.nav_month = nav_m + 1
            st.rerun()

    if nav_y != today.year or nav_m != today.month:
        if st.button("↩ Về tháng này", key="nav_today", use_container_width=False):
            st.session_state.nav_month = today.month
            st.session_state.nav_year  = today.year
            st.rerun()

    with st.expander("⚙️ Kỳ xem mở rộng (Quý/Năm/Tùy chỉnh)", expanded=False):
        ex1, ex2 = st.columns(2)
        with ex1: ky2_ext = st.selectbox("Kỳ", PERIOD_PRESETS, key="ky2_ext")
        with ex2: nam2_ext = st.selectbox("Năm", list(range(2023, today.year + 2)),
                                           index=list(range(2023, today.year + 2)).index(today.year), key="nam2_ext")
        cs2 = ce2 = None
        if ky2_ext == "Tùy chỉnh khoảng ngày":
            cx1, cx2 = st.columns(2)
            with cx1: cs2 = st.date_input("Từ ngày", date(nam2_ext, 1, 1), key="cs2")
            with cx2: ce2 = st.date_input("Đến ngày", today,               key="ce2")
        s2_ext, e2_ext = get_date_range(ky2_ext, nam2_ext, cs2, ce2)
        use_ext = st.checkbox("Dùng kỳ mở rộng này thay vì điều hướng tháng", key="use_ext")

    if "use_ext" in st.session_state and st.session_state.use_ext:
        s2 = s2_ext; e2 = e2_ext
    else:
        s2 = date(nav_y, nav_m, 1)
        e2 = date(nav_y, nav_m, calendar.monthrange(nav_y, nav_m)[1])

    sd2_dau, t2_thu, t2_chi, sd2_cuoi = compute_balances(df_real2, s2, e2)

    st.markdown(f"""
    <div class="sb">
        <div class="sc dau">
            <div class="sl">SỐ DƯ ĐẦU KỲ</div>
            <div class="sv" style="color:{'#1a9e5f' if sd2_dau>=0 else '#e05252'}">{fmt_vnd(sd2_dau)}</div>
        </div>
        <div class="sc thu">
            <div class="sl">TỔNG THU</div>
            <div class="sv">+{fmt_vnd(t2_thu)}</div>
        </div>
        <div class="sc chi">
            <div class="sl">TỔNG CHI</div>
            <div class="sv">-{fmt_vnd(t2_chi)}</div>
        </div>
        <div class="sc end">
            <div class="sl">SỐ DƯ CUỐI KỲ</div>
            <div class="sv">{fmt_vnd(sd2_cuoi)}</div>
        </div>
    </div>
    """, unsafe_allow_html=True)

    # Pending warning badge
    df_cho_tt = df_all2[df_all2["trang_thai"] == "Chờ thanh toán"] if not df_all2.empty else pd.DataFrame()
    if not df_cho_tt.empty:
        tong_cho = df_cho_tt["so_tien"].sum()
        st.markdown(f"""
        <div class="pending-badge">
            ⏳ <b>{len(df_cho_tt)} giao dịch</b> · <b>{fmt_vnd(tong_cho)}</b>
            đang <b>Chờ thanh toán</b> — chưa tính vào sổ này.
            Vào tab <b>📤 Import & Thu Tiền</b> để xác nhận.
        </div>
        """, unsafe_allow_html=True)

    df_f2 = df_real2.copy() if not df_real2.empty else pd.DataFrame()
    if not df_f2.empty:
        df_f2 = df_f2[(df_f2["ngay"].dt.date >= s2) & (df_f2["ngay"].dt.date <= e2)]
        if phong2 != "Tất cả": df_f2 = df_f2[df_f2["ma_phong"] == phong2]
        if loai2  != "Tất cả": df_f2 = df_f2[df_f2["loai"] == loai2]
        df_f2 = df_f2.sort_values("ngay")

    if df_f2.empty:
        st.info("Không có giao dịch thực thu trong kỳ đã chọn.")
    else:
        df_f2["ngay_str"] = df_f2["ngay"].dt.strftime("%d/%m/%Y")
        for nstr, grp in df_f2.groupby("ngay_str", sort=False):
            net = grp[grp["loai"] == "Thu"]["so_tien"].sum() - grp[grp["loai"] == "Chi"]["so_tien"].sum()
            nc  = "#1a9e5f" if net >= 0 else "#e05252"
            st.markdown(f"""
            <div style="display:flex;justify-content:space-between;padding:5px 4px;
                        border-bottom:1px solid #f0f0f0;margin:10px 0 4px">
                <div style="font-weight:700;color:#333">{nstr}</div>
                <div style="font-size:.84rem;font-weight:700;color:{nc}">{'+'if net>=0 else ''}{fmt_vnd(net)}</div>
            </div>
            """, unsafe_allow_html=True)
            for _, r in grp.iterrows():
                icon = "📈" if r["loai"] == "Thu" else "📉"
                cls  = "thu" if r["loai"] == "Thu" else "chi"
                pref = "+" if r["loai"] == "Thu" else "-"
                note = f" · {r['ghi_chu']}" if str(r.get("ghi_chu", "")).strip() else ""
                gd_id   = int(r["id"])
                is_sel  = st.session_state.selected_gd == gd_id
                is_edit = st.session_state.edit_id     == gd_id

                col_card, col_tap = st.columns([20, 1])
                with col_card:
                    st.markdown(f"""
                    <div class="tx">
                      <div class="ti {cls}">{icon}</div>
                      <div class="tn">
                        <div class="tc">{r['danh_muc']}</div>
                        <div class="tm">{r['ma_phong']}{note}</div>
                      </div>
                      <div class="ta {cls}">{pref}{fmt_vnd(r['so_tien'])}</div>
                    </div>
                    """, unsafe_allow_html=True)
                with col_tap:
                    tap_label = "▲" if is_sel else "▼"
                    if st.button(tap_label, key=f"tap_{gd_id}", use_container_width=True):
                        if is_sel:
                            st.session_state.selected_gd = None
                            st.session_state.edit_id     = None
                        else:
                            st.session_state.selected_gd = gd_id
                            st.session_state.edit_id     = None
                        st.rerun()

                # ── Action bar ───────────────────────────────────────────────
                if is_sel and not is_edit:
                    ab1, ab2, ab3 = st.columns(3)
                    with ab1:
                        if st.button("✏️  Sửa", key=f"act_edit_{gd_id}",
                                     use_container_width=True, type="primary"):
                            st.session_state.edit_id = gd_id
                            st.rerun()
                    with ab2:
                        if st.button("📋  Copy", key=f"act_copy_{gd_id}",
                                     use_container_width=True):
                            insert_gd(r["ngay"].date(), r["loai"], r["ma_phong"],
                                      r["danh_muc"], int(r["so_tien"]),
                                      str(r.get("ghi_chu", "") or ""),
                                      trang_thai="Đã thu")
                            st.session_state.selected_gd = None
                            st.rerun()
                    with ab3:
                        if st.button("🗑️  Xóa", key=f"act_del_{gd_id}",
                                     use_container_width=True):
                            delete_gd(gd_id)
                            st.session_state.selected_gd = None
                            st.rerun()

                # ── Form sửa inline ──────────────────────────────────────────
                if is_edit:
                    st.markdown('<div class="edit-form"><div class="edit-form-title">✏️ Sửa giao dịch</div>', unsafe_allow_html=True)
                    ef1, ef2 = st.columns(2)
                    with ef1:
                        e_loai = st.radio("Loại", ["Thu", "Chi"],
                            index=0 if r["loai"] == "Thu" else 1,
                            horizontal=True, key=f"eloai_{gd_id}")
                        e_ngay = st.date_input("Ngày", value=r["ngay"].date(), key=f"engay_{gd_id}")
                        if e_loai == "Thu":
                            e_phong = st.selectbox("Phòng", ROOMS,
                                index=ROOMS.index(r["ma_phong"]) if r["ma_phong"] in ROOMS else 0,
                                key=f"ephong_{gd_id}")
                            e_dm = st.selectbox("Danh mục", DANH_MUC_THU,
                                index=DANH_MUC_THU.index(r["danh_muc"]) if r["danh_muc"] in DANH_MUC_THU else 0,
                                key=f"edm_{gd_id}")
                        else:
                            e_phong = "Chung"
                            e_dm = st.selectbox("Danh mục", DANH_MUC_CHI,
                                index=DANH_MUC_CHI.index(r["danh_muc"]) if r["danh_muc"] in DANH_MUC_CHI else 0,
                                key=f"edm_{gd_id}")
                    with ef2:
                        e_tien_raw = st.text_input("Số tiền (₫)", value=str(int(r["so_tien"])),
                            placeholder="VD: 1500000", key=f"etien_{gd_id}")
                        e_tien = parse_tien(e_tien_raw)
                        e_gc = st.text_input("Ghi chú", value=str(r.get("ghi_chu", "") or ""),
                            key=f"egc_{gd_id}")
                        st.markdown("<br>", unsafe_allow_html=True)
                        btn_luu, btn_huy = st.columns(2)
                        with btn_luu:
                            if st.button("💾 Lưu", key=f"esave_{gd_id}", type="primary"):
                                ok = update_gd(gd_id, e_ngay, e_loai,
                                               e_phong if e_loai == "Thu" else "Chung",
                                               e_dm, e_tien, e_gc)
                                if ok:
                                    st.toast("✅ Đã cập nhật giao dịch!", icon="💾")
                                    st.session_state.edit_id     = None
                                    st.session_state.selected_gd = None
                                    st.rerun()
                        with btn_huy:
                            if st.button("✖ Huỷ", key=f"ecancel_{gd_id}"):
                                st.session_state.edit_id     = None
                                st.session_state.selected_gd = None
                                st.rerun()
                    st.markdown('</div>', unsafe_allow_html=True)

    st.markdown("---")
    if not df_real2.empty:
        st.download_button(
            f"📥 Xuất Excel Dòng Tiền ({s2.strftime('%d/%m/%Y')} – {e2.strftime('%d/%m/%Y')})",
            data=xuat_dong_tien(df_real2, s2, e2),
            file_name=f"DongTien_{s2.strftime('%d%m%Y')}_{e2.strftime('%d%m%Y')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )


# ════════════════════════════════════════════════════════════════════════════
# TAB 4 — XUẤT SỔ THUẾ
# ════════════════════════════════════════════════════════════════════════════
with tab4:
    df_all3 = load_all()
    df_real3 = df_all3[df_all3["trang_thai"].isin(TRANG_THAI_DA_THU)].copy() if not df_all3.empty else df_all3.copy()

    st.markdown('<div class="section-heading"><span>📊</span> Xuất Sổ Thuế — TT152/2025/TT-BTC</div>', unsafe_allow_html=True)
    st.markdown("""<div class="tax-info">
        <b>📋 S1a-HKD</b>: Doanh thu tiền phòng (không bao gồm điện/nước thu hộ).<br>
        <b>📋 S2a-HKD</b>: Chỉ tiền phòng · Tự tính <b>GTGT 5%</b> + <b>TNCN 2%</b>.<br>
        <b>⚠️ Lưu ý:</b> Chỉ tính giao dịch trạng thái <b>Đã thu / Đã thanh toán</b>.
    </div>""", unsafe_allow_html=True)

    tc1, tc2 = st.columns([3, 1])
    with tc1: ky3  = st.selectbox("📅 Kỳ kê khai", KY_OPTIONS, key="ky3")
    with tc2: nam3 = st.selectbox("Năm", list(range(2023, today.year + 2)),
                                   index=list(range(2023, today.year + 2)).index(today.year), key="nam3")
    cs3 = ce3 = None
    if ky3 == "Tùy chỉnh khoảng ngày":
        cx3a, cx3b = st.columns(2)
        with cx3a: cs3 = st.date_input("Từ ngày", date(nam3, 1, 1), key="cs3")
        with cx3b: ce3 = st.date_input("Đến ngày", today,           key="ce3")

    s3, e3 = get_date_range(ky3, nam3, cs3, ce3)
    st.markdown(f"<div style='font-size:.84rem;color:#888;margin:4px 0 14px'>📆 Kỳ: <b>{s3.strftime('%d/%m/%Y')}</b> → <b>{e3.strftime('%d/%m/%Y')}</b></div>", unsafe_allow_html=True)

    if not df_real3.empty:
        df_ky3  = df_real3[(df_real3["ngay"].dt.date >= s3) & (df_real3["ngay"].dt.date <= e3)]
        n_thu   = len(df_ky3[(df_ky3["loai"] == "Thu") & (df_ky3["danh_muc"] == "Doanh thu tiền phòng")])
        n_phong = n_thu
        tong_s1 = df_ky3[(df_ky3["loai"] == "Thu") & (df_ky3["danh_muc"] == "Doanh thu tiền phòng")]["so_tien"].sum()
        tong_s2 = tong_s1

        m1, m2, m3, m4 = st.columns(4)
        m1.metric("Giao dịch Thu (S1a)", n_thu)
        m2.metric("Giao dịch phòng (S2a)", n_phong)
        m3.metric("Tổng Thu (S1a)", fmt_vnd(tong_s1))
        m4.metric("Doanh thu phòng (S2a)", fmt_vnd(tong_s2))

        if tong_s2 > 0:
            st.markdown(f"""
            <div style="background:#fff5f0;border-radius:12px;padding:14px 20px;margin:12px 0;
                        font-size:.88rem;display:flex;gap:32px;flex-wrap:wrap;border:1.5px solid #fdd;">
                <div>🟡 <b>GTGT (5%)</b>: <span style="font-weight:800;color:#c0392b;font-size:1.05rem">{fmt_vnd(round(tong_s2*0.05))}</span></div>
                <div>🔵 <b>TNCN (2%)</b>: <span style="font-weight:800;color:#c0392b;font-size:1.05rem">{fmt_vnd(round(tong_s2*0.02))}</span></div>
                <div>💰 <b>Tổng thuế (7%)</b>: <span style="font-weight:800;color:#8b0000;font-size:1.05rem">{fmt_vnd(round(tong_s2*0.07))}</span></div>
            </div>
            """, unsafe_allow_html=True)

        st.markdown("---")
        xc1, xc2 = st.columns(2)
        with xc1:
            st.markdown("**📗 Sổ S1a-HKD — Doanh thu tiền phòng**")
            st.markdown("<div style='font-size:.79rem;color:#888;margin-bottom:10px'>Chỉ doanh thu tiền phòng · Không tính điện/nước thu hộ</div>", unsafe_allow_html=True)
            if n_thu > 0:
                st.download_button("📥 Xuất S1a-HKD",
                    data=xuat_s1a(df_real3, s3, e3),
                    file_name=f"S1a_HKD_{s3.strftime('%d%m%Y')}_{e3.strftime('%d%m%Y')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True)
            else:
                st.info("Không có dữ liệu Thu thực tế trong kỳ.")
        with xc2:
            st.markdown("**📘 Sổ S2a-HKD — Phân loại + Tính thuế**")
            st.markdown("<div style='font-size:.79rem;color:#888;margin-bottom:10px'>Chỉ tiền phòng · GTGT 5% + TNCN 2%</div>", unsafe_allow_html=True)
            if n_phong > 0:
                st.download_button("📥 Xuất S2a-HKD",
                    data=xuat_s2a(df_real3, s3, e3),
                    file_name=f"S2a_HKD_{s3.strftime('%d%m%Y')}_{e3.strftime('%d%m%Y')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True)
            else:
                st.info("Không có doanh thu tiền phòng thực tế trong kỳ.")
    else:
        st.info("Chưa có dữ liệu thực thu.")

# ════════════════════════════════════════════════════════════════════════════
# TAB 5 — QUẢN LÝ TIỀN CỌC
# Dữ liệu cọc TUYỆT ĐỐI tách biệt — không lọt vào thu chi / sổ thuế
# ════════════════════════════════════════════════════════════════════════════
with tab5:

    # ── Helper: lấy giá phòng hiện tại từ dữ liệu giao_dich (tham chiếu ngầm)
    def _get_gia_phong_hien_tai() -> dict:
        """
        Lấy giá phòng (tiền thuê phòng) mới nhất của mỗi phòng từ tab giao_dich.
        Dùng để đối chiếu ngầm với tiền cọc — chỉ đọc, không ảnh hưởng logic.
        """
        try:
            df_ref = load_all()
            if df_ref.empty:
                return {}
            df_phong = df_ref[
                (df_ref["loai"] == "Thu") &
                (df_ref["danh_muc"] == "Doanh thu tiền phòng") &
                (df_ref["trang_thai"].isin(TRANG_THAI_DA_THU))
            ].copy()
            if df_phong.empty:
                return {}
            # Lấy giao dịch mới nhất của mỗi phòng
            idx_latest = df_phong.groupby("ma_phong")["ngay"].idxmax()
            df_latest  = df_phong.loc[idx_latest]
            return dict(zip(df_latest["ma_phong"], df_latest["so_tien"]))
        except Exception:
            return {}

    st.markdown('<div class="section-heading"><span>📌</span> Quản lý Tiền Cọc — 9 Phòng</div>', unsafe_allow_html=True)
    st.markdown("""
    <div style="background:linear-gradient(135deg,#f4f0ff,#ede8ff);border:1.5px solid #c9baff;
                border-radius:12px;padding:12px 18px;margin-bottom:18px;font-size:.86rem;color:#5b2dd9;">
        <b>📌 Khoản giữ hộ:</b> Tiền cọc <b>không được</b> tính vào dòng tiền, báo cáo thu chi,
        hay sổ thuế S1a/S2a. Đây là khu vực quản lý riêng biệt, chỉ để theo dõi số tiền đang giữ hộ khách thuê.
    </div>
    """, unsafe_allow_html=True)

    # ── Tải dữ liệu cọc + giá phòng tham chiếu
    df_coc     = load_coc()
    gia_phong  = _get_gia_phong_hien_tai()

    # ═══ LAYOUT: Trên (bảng tổng hợp) / Dưới (form cập nhật fragment) ════════
    # ── PHẦN TRÊN: Bảng tổng hợp + Tổng quỹ cọc ─────────────────────────────
    st.markdown("#### 📋 Bảng Tổng Hợp Tiền Cọc Hiện Tại")

    tong_coc = int(df_coc["so_tien_coc"].sum())
    n_co_coc = len(df_coc[df_coc["so_tien_coc"] > 0])

    # Thẻ tổng quỹ cọc
    st.markdown(f"""
    <div class="coc-total-box">
        <div class="coc-total-label">Tổng Quỹ Tiền Cọc Đang Giữ Hộ</div>
        <div class="coc-total-amount">{fmt_vnd(tong_coc)}</div>
        <div class="coc-total-sub">{n_co_coc} / {len(ROOMS)} phòng đang có cọc · Đây là khoản giữ hộ, sẽ trả lại khách khi trả phòng</div>
    </div>
    """, unsafe_allow_html=True)

    # Hiển thị lưới 9 phòng — render từng hàng 3 phòng để đảm bảo thứ tự
    # P101 P201 P202 / P301 P302 P401 / P402 P501 P502 — đúng trên cả mobile lẫn desktop
    rooms_data = []
    for i, row in df_coc.iterrows():
        phong    = row["ma_phong"]
        ten_kh   = str(row.get("ten_kh", "") or "").strip()
        so_coc   = int(row.get("so_tien_coc", 0) or 0)
        ngay_coc = row.get("ngay_coc")
        ghi_chu  = str(row.get("ghi_chu", "") or "").strip()
        gia_ht   = gia_phong.get(phong, 0)
        rooms_data.append((phong, ten_kh, so_coc, ngay_coc, ghi_chu, gia_ht))

    # Render từng hàng 3 cột — thứ tự luôn đúng vì Python loop theo hàng
    for row_start in range(0, len(rooms_data), 3):
        row_items = rooms_data[row_start : row_start + 3]
        cols = st.columns(len(row_items))
        for col, (phong, ten_kh, so_coc, ngay_coc, ghi_chu, gia_ht) in zip(cols, row_items):
            warn_html = ""
            if so_coc > 0 and gia_ht > 0 and so_coc < gia_ht:
                warn_html = f'<div class="coc-warn">⚠️ Cọc &lt; Giá phòng hiện tại ({fmt_vnd(gia_ht)})</div>'
            if so_coc > 0:
                ngay_str    = ngay_coc.strftime("%d/%m/%Y") if pd.notna(ngay_coc) else "—"
                meta_parts  = [f"📅 {ngay_str}"]
                if ten_kh:  meta_parts.append(f"👤 {ten_kh}")
                if ghi_chu: meta_parts.append(f"📝 {ghi_chu}")
                meta = " · ".join(meta_parts)
                card = f"""<div class="coc-card">
  <div class="coc-header">📦 {phong}</div>
  <div class="coc-room">{phong}</div>
  <div class="coc-amount">{fmt_vnd(so_coc)}</div>
  <div class="coc-meta">{meta}</div>
  {warn_html}
</div>"""
            else:
                card = f"""<div class="coc-card coc-empty">
  <div class="coc-header">📦 {phong}</div>
  <div class="coc-room">{phong}</div>
  <div class="coc-amount" style="color:#bbb">Chưa có cọc</div>
  <div class="coc-meta">— Chưa cập nhật —</div>
</div>"""
            with col:
                st.markdown(card, unsafe_allow_html=True)

    st.markdown("---")

    # ═══ PHẦN DƯỚI: Form cập nhật (dùng @st.fragment để tránh lag) ════════════
    @st.fragment
    def fragment_cap_nhat_coc():
        cfv = st.session_state.coc_form_ver

        st.markdown('<div class="section-heading"><span>✏️</span> Cập Nhật Thông Tin Tiền Cọc</div>', unsafe_allow_html=True)

        fc1, fc2 = st.columns([1, 2], gap="large")

        # ── Cột trái: chọn phòng + preview cọc hiện tại ──────────────────────
        with fc1:
            phong_sel = st.selectbox(
                "🏠 Chọn phòng",
                ROOMS,
                key=f"coc_phong_sel_{cfv}",
            )
            df_coc_now = load_coc()
            row_now    = df_coc_now[df_coc_now["ma_phong"] == phong_sel]
            if not row_now.empty:
                r          = row_now.iloc[0]
                so_coc_now = int(r.get("so_tien_coc", 0) or 0)
                if so_coc_now > 0:
                    ten_now   = str(r.get("ten_kh",  "") or "")
                    ngay_now  = r.get("ngay_coc")
                    ngay_disp = ngay_now.strftime("%d/%m/%Y") if pd.notna(ngay_now) else "—"
                    st.markdown(f"""
                    <div style="background:#f4f0ff;border:1px solid #c9baff;border-radius:10px;
                                padding:12px 14px;font-size:.83rem;color:#5b2dd9;">
                        <b>📌 Đang giữ cọc {phong_sel}:</b><br>
                        💰 {fmt_vnd(so_coc_now)}<br>
                        👤 {ten_now or '—'}<br>
                        📅 {ngay_disp}
                    </div>
                    """, unsafe_allow_html=True)
                else:
                    st.markdown(f"""
                    <div style="background:#f9f9f9;border:1px dashed #ddd;border-radius:10px;
                                padding:12px 14px;font-size:.83rem;color:#aaa;">
                        <i>Phòng {phong_sel} chưa có tiền cọc.</i>
                    </div>
                    """, unsafe_allow_html=True)

        # ── Cột phải: form nhập liệu (Enter = submit) ─────────────────────────
        with fc2:
            # Tên khách + ngày cọc bên ngoài form (không cần Enter submit)
            ten_kh_inp   = st.text_input(
                "👤 Tên khách thuê",
                placeholder="Nguyễn Văn A",
                key=f"coc_ten_kh_{cfv}",
            )
            ngay_coc_inp = st.date_input(
                "📅 Ngày đặt cọc",
                value=today,
                key=f"coc_ngay_{cfv}",
            )

            # Số tiền bên ngoài form để hiển thị preview realtime
            so_coc_raw = st.text_input(
                "💰 Số tiền cọc (₫)",
                placeholder="VD: 4.500.000",
                key=f"coc_so_tien_{cfv}",
                on_change=lambda: st.session_state.update(
                    coc_live_tien=st.session_state.get(f"coc_so_tien_{cfv}", "")
                ),
            )
            so_coc_val = parse_tien(so_coc_raw)
            if so_coc_val > 0:
                st.markdown(
                    f"<div style='color:#5b2dd9;font-weight:800;font-size:1.2rem;"
                    f"margin:-2px 0 6px'>{fmt_vnd(so_coc_val)}</div>",
                    unsafe_allow_html=True,
                )

            # Ghi chú + nút Lưu cùng trong 1 st.form → Enter sẽ submit
            with st.form(key=f"form_coc_{cfv}", clear_on_submit=True):
                ghi_chu_inp   = st.text_input(
                    "📝 Ghi chú (tùy chọn)",
                    placeholder="VD: Cọc hợp đồng T01/2025, ...",
                    key=f"coc_ghichu_{cfv}",
                )
                submitted_coc = st.form_submit_button(
                    "💾  Lưu thông tin cọc",
                    type="primary",
                    use_container_width=True,
                )

            if submitted_coc:
                so_coc_submit = parse_tien(
                    st.session_state.get(f"coc_so_tien_{cfv}", "")
                )
                ten_submit = st.session_state.get(f"coc_ten_kh_{cfv}", "").strip()
                if so_coc_submit <= 0:
                    st.error("⚠️ Vui lòng nhập số tiền cọc lớn hơn 0")
                elif not ten_submit:
                    st.error("⚠️ Vui lòng nhập tên khách thuê")
                else:
                    with st.spinner(f"Đang lưu tiền cọc {phong_sel}..."):
                        ok = upsert_coc(
                            ma_phong    = phong_sel,
                            ten_kh      = ten_submit,
                            ngay_coc    = st.session_state.get(f"coc_ngay_{cfv}", today),
                            so_tien_coc = so_coc_submit,
                            ghi_chu     = ghi_chu_inp.strip(),
                        )
                    if ok:
                        st.toast(
                            f"✅ Đã lưu cọc {phong_sel} · {ten_submit} · {fmt_vnd(so_coc_submit)}",
                            icon="📌",
                        )
                        st.session_state.coc_form_ver += 1
                        st.session_state.coc_live_tien = ""
                        st.rerun(scope="fragment")
                    else:
                        st.error("❌ Lỗi khi lưu. Kiểm tra kết nối Google Sheets.")

    fragment_cap_nhat_coc()
